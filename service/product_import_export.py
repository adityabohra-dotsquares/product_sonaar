from typing import List, Optional, Tuple
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from models.product import Product, ProductVariant
from datetime import datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import or_, func, and_, update
from models.b_tasks import BackgroundTask

# from apis.v1.import_export import EXPECTED_HEADERS, build_category_info
from models.category import Category
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from models.brand import Brand


async def build_category_info(
    db: AsyncSession, category_id: str
) -> Tuple[str, str, str]:
    """
    Returns (category_identifier, breadcrumb, product_type)

    - category_identifier: category_code if present, otherwise the category ID
    - breadcrumb: "Root > Parent > ... > Leaf"
    - product_type: leaf category name
    """

    result = await db.execute(select(Category).where(Category.id == category_id))
    category: Category | None = result.scalar_one_or_none()

    if not category:
        # Category id does not exist, return id + empty strings
        return category_id, "", ""

    names: List[str] = []
    current: Category | None = category

    # Safety counter in case of bad data (cycles)
    max_depth = 20
    depth = 0

    while current and depth < max_depth:
        names.append(current.name)

        # If there is no parent_id, we've reached the root
        if current.parent_id is None:
            break

        # Load parent from DB using parent_id
        parent_result = await db.execute(
            select(Category).where(Category.id == current.parent_id)
        )
        current = parent_result.scalar_one_or_none()
        depth += 1

        if current is None:
            # Broken parent_id, stop here
            break

    # We collected from leaf → root, so reverse to get root → leaf
    names.reverse()

    breadcrumb = " > ".join(names)
    product_type = category.name  # leaf name

    # Key change: prefer category_code if available
    category_identifier = category.category_code or str(category.id)

    return category_identifier, breadcrumb, product_type


EXPECTED_HEADERS = [
    "category code",
    "brand code",
    "SKU",
    "Product Code",
    "Title",
    "Long Description",
    "Brand Name",
    "Condition",
    "Product Status",
    "Supplier Name",
    "Product Tag",
    "Parent SKU / Variation Group Code",
    "Variant Option 1 (e.g., Colour, Size, Style)",
    "Variant Values 1",
    "Variant Option 2 (e.g., Colour, Size, Style)",
    "Variant Values 2",
    "Bundle / Combo Indicator",
    "Selling Price",
    "RRP",
    "Cost Price",
    "Stock Quantity",
    "Package Weight (kg)",
    "Package Length(Cms)",
    "Package Width(Cms)",
    "Package Height(Cms)",
    "Harmonized Code (HS Code)",
    "Shipping Template",
    "Handling Time (days)",
    "Fast Dispatch",
    "Free Shipping",
    "Ships From Location",
    "Product Video Link",
    "Main Image URL",
    "Lifestyle Image URL",
    "Image 1 URL",
    "Image 2 URL",
    "Image 3 URL",
    "Image 4 URL",
    "Image 5 URL",
    "Image 6 URL",
    "Image 7 URL",
    "Image 8 URL",
    "Image 9 URL",
    "Image 10 URL",
    "Precautionary Note",
    "Care Instructions (If Applicable)",
    "Warranty / Guarantee (If Applicable)",
    "SEO Keywords",
    "Page Title",
    "Meta Description",
    "URL Handles",
    "Canonical URL",
]


def build_sqlalchemy_filters(filters_dict: dict):
    filters = []

    if not filters_dict:
        return filters

    if status := filters_dict.get("status"):
        filters.append(Product.status == status)

    if vendor := filters_dict.get("vendor"):
        filters.append(Product.vendor_id.ilike(f"%{vendor}%"))

    if brand_id := filters_dict.get("brand_id"):
        filters.append(Product.brand.has(Brand.name.ilike(f"%{brand_id}%")))

    if ships_from := filters_dict.get("ships_from"):
        filters.append(Product.ships_from_location.ilike(f"%{ships_from}%"))

    if tag := filters_dict.get("tag"):
        tags = [t.strip() for t in tag.split(",") if t.strip()]
        if tags:
            filters.append(or_(*[Product.tags.contains([t]) for t in tags]))

    if title := filters_dict.get("title"):
        filters.append(Product.title.ilike(f"%{title}%"))

    if sku := filters_dict.get("sku"):
        filters.append(Product.sku == sku)

    if product_type := filters_dict.get("product_type"):
        filters.append(Product.product_type.ilike(f"%{product_type}%"))

    if filters_dict.get("fast_dispatch") is not None:
        filters.append(Product.fast_dispatch == filters_dict["fast_dispatch"])

    if filters_dict.get("free_shipping") is not None:
        filters.append(Product.free_shipping == filters_dict["free_shipping"])

    if q := filters_dict.get("q"):
        q_lower = q.lower()
        filters.append(
            or_(
                func.lower(Product.title).like(f"%{q_lower}%"),
                func.lower(Product.sku).like(f"%{q_lower}%"),
            )
        )

    if ean := filters_dict.get("ean"):
        filters.append(Product.ean == ean)

    if asin := filters_dict.get("asin"):
        filters.append(Product.asin == asin)

    if mpn := filters_dict.get("mpn"):
        filters.append(Product.mpn == mpn)

    if condition := filters_dict.get("condition"):
        filters.append(Product.product_condition == condition)

    return filters


# -------------------------------------------------------
# 1️⃣ CORE — shared export logic (NO files / HTTP here)
# -------------------------------------------------------
async def build_export_data(
    db: AsyncSession,
    product_ids,
    filters: dict,
    sort_key: str,
    columns: Optional[List[str]] = None,
    task_id: Optional[str] = None,
):
    print("BUILDING EXPORT DATA")
    query = select(Product).options(
        selectinload(Product.category),
        selectinload(Product.images),
        selectinload(Product.variants).selectinload(ProductVariant.attributes),
        selectinload(Product.variants).selectinload(ProductVariant.images),
        selectinload(Product.brand),
        selectinload(Product.seo),
    )
    print("QUERY", query)
    if product_ids:
        query = query.where(Product.id.in_(product_ids))

    sqlalchemy_filters = build_sqlalchemy_filters(filters)

    if sqlalchemy_filters:
        query = query.where(and_(*sqlalchemy_filters))

    SORT_MAP = {
        "price_asc": Product.price.asc(),
        "price_desc": Product.price.desc(),
        "newly_added": Product.created_at.desc(),
        "oldest": Product.created_at.asc(),
        "stock_asc": Product.stock.asc(),
        "stock_desc": Product.stock.desc(),
    }

    if sort_key:
        sort_clause = SORT_MAP.get(sort_key)
        if sort_clause is None:
            raise HTTPException(status_code=400, detail=f"Invalid sort key: {sort_key}")
        query = query.order_by(sort_clause)
    else:
        query = query.order_by(Product.created_at.desc())

    result = await db.execute(query)
    products = result.scalars().unique().all()
    

    # Collect all unique category_ids and all parent_ids
    category_ids = {p.category_id for p in products if p.category_id}

    # Fetch all categories at once
    result = await db.execute(
        select(Category).where(Category.id.in_(category_ids))
    )
    categories = result.scalars().all()

    # Build a map of category_id → Category
    category_map = {c.id: c for c in categories}
    print("CATEGORY MAP", category_map)
    result = await db.execute(select(Category))
    all_categories = result.scalars().all()
    all_category_map = {c.id: c for c in all_categories}
    print("ALL CATEGORY MAP", all_category_map)
    print("PRODUCTS", products)
    def build_breadcrumb(category_id: str, category_map: dict) -> Tuple[str, str, str]:
        """
        Returns (category_identifier, breadcrumb, product_type)
        """
        current = category_map.get(category_id)
        if not current:
            return category_id, "", ""

        names = []
        depth = 0
        max_depth = 20

        while current and depth < max_depth:
            names.append(current.name)
            if not current.parent_id:
                break
            current = category_map.get(current.parent_id)
            depth += 1

        names.reverse()
        breadcrumb = " > ".join(names)
        product_type = category_map[category_id].name
        category_identifier = category_map[category_id].category_code or str(category_id)

        return category_identifier, breadcrumb, product_type

    # exit()




    
    rows = []

    # ---------------- helpers ----------------
    def _dec_or_empty(val):
        return "" if val is None else str(val)

    def _join_tags(tags):
        return "" if not tags else ", ".join(str(t) for t in tags)

    def _collect_image_urls(product, variant):
        urls = []
        if variant and variant.images:
            main = [i.image_url for i in variant.images if i.is_main and i.image_url]
            others = [
                i.image_url for i in variant.images if not i.is_main and i.image_url
            ]
            urls.extend(main + others)

        if product.images:
            main = [i.image_url for i in product.images if i.is_main and i.image_url]
            others = [
                i.image_url for i in product.images if not i.is_main and i.image_url
            ]
            urls.extend(main + others)

        seen, ordered = set(), []
        for u in urls:
            if u not in seen:
                seen.add(u)
                ordered.append(u)
        return ordered

    def _collect_video_url(product, variant):
        if variant and variant.images:
            for i in variant.images:
                if i.video_url:
                    return i.video_url
        if product.images:
            for i in product.images:
                if i.video_url:
                    return i.video_url
        return ""
    print("AFTER PRODUCTS UPDATEDDDDD")
    # Build rows
    processed_count = 0
    total_products = len(products)
    if task_id:
        await db.execute(
            update(BackgroundTask)
            .where(BackgroundTask.task_id == task_id)
            .values(task_info={"total_rows": total_products, "processed_rows": processed_count})
        )
        await db.commit()

    for product in products:
        # Category info (first 3 columns)
        if product.category_id:
            category_code, category_breadcrumb, leaf_product_type = (
                build_breadcrumb(product.category_id,all_category_map)
            )
        else:
            category_code, category_breadcrumb, leaf_product_type = "", "", ""

        # base fields used for both product & variants
        product_brand_code = product.ean or ""  # change if Brand has separate 'code'
        product_code = product.product_code or ""
        supplier_name = product.supplier or ""
        product_tag = _join_tags(product.tags)
        hs_code = product.hs_code or ""
        shipping_template = product.shipping_template or ""
        precautionary_note = product.precautionary_note or ""
        care_instructions = product.care_instructions or ""
        warranty = product.warranty or ""
        
        # SEO fields from ProductSEO
        seo_keywords = ""
        page_title = ""
        meta_description = ""
        url_handle = ""
        canonical_url = ""
        
        if product.seo:
            seo_keywords = product.seo.meta_keywords or ""
            page_title = product.seo.page_title or ""
            meta_description = product.seo.meta_description or ""
            url_handle = product.seo.url_handle or ""
            canonical_url = product.seo.canonical_url or ""

        cost_price = product.cost_price or ""

        # Get all images & video at product level (fallback)
        product_level_imgs = _collect_image_urls(product, None)
        product_level_video = _collect_video_url(product, None)

        # Prepare parent row (always needed)
        selling_price = product.price
        rrp = product.rrp_price
        cost_price = product.cost_price
        stock_qty = product.stock

        weight = product.weight
        length = product.length
        width = product.width
        height = product.height
        handling_time = product.handling_time_days
        ships_from_location = product.ships_from_location or ""

        img_urls = product_level_imgs
        video_url = product_level_video

        main_image = img_urls[0] if len(img_urls) >= 1 else ""
        lifestyle_image = img_urls[1] if len(img_urls) >= 2 else ""
        extra_imgs = img_urls[2:12]
        extra_imgs += [""] * (10 - len(extra_imgs))

        parent_row = [
            # 1–3: category info
            category_code or "",
            category_breadcrumb or "",
            leaf_product_type or "",
            # 5. "SKU"
            product.sku or "",
            # 6. "Product Code"
            product_code,
            # 7. "Title"
            product.title or "",
            # 8. "Long Description"
            product.description or "",
            # 9. "Brand_id"
            product.brand.name if product.brand else "",
            # 10. "Condition"
            product.product_condition or "",
            # 11. "Product Status"
            product.status or "",
            # 12. "Supplier Name"
            supplier_name,
            # 13. "Product Tag"
            product_tag,
            # 14. "Parent SKU / Variation Group Code"
            "",
            # 15–18: no variant attributes for parent row
            "",
            "",
            "",
            "",
            # 19. "Bundle / Combo Indicator"
            product.bundle_group_code or "",
            # 20. "Selling Price"
            _dec_or_empty(selling_price),
            # 21. "RRP"
            _dec_or_empty(rrp),
            _dec_or_empty(cost_price),
            # 22. "Stock Quantity"
            str(stock_qty or ""),
            # 23–26: package dimensions
            _dec_or_empty(weight),
            _dec_or_empty(length),
            _dec_or_empty(width),
            _dec_or_empty(height),
            # 27. "Harmonized Code (HS Code)"
            hs_code,
            # 28. "Shipping Template"
            shipping_template,
            # 29. "Handling Time (days)"
            str(handling_time or "") if handling_time is not None else "",
            # 30. "Fast Dispatch"
            product.fast_dispatch if product.fast_dispatch is not None else "",
            # 31. "Free Shipping"
            product.free_shipping if product.free_shipping is not None else "",
            # 32. "Ships From Location"
            ships_from_location,
            # 33. "Product Video Link"
            video_url,
            # 34. "Main Image URL"
            main_image,
            # 35. "Lifestyle Image URL"
            lifestyle_image,
            # 36–45: images
            *extra_imgs,
            # 46. "Precautionary Note"
            precautionary_note,
            # 46. "Care Instructions (If Applicable)"
            care_instructions,
            # 47. "Warranty / Guarantee (If Applicable)"
            warranty,
            # 48. "SEO Keywords"
            seo_keywords,
            # 49. "Page Title"
            page_title,
            # 50. "Meta Description"
            meta_description,
            # 51. "URL Handles"
            url_handle,
            # 52. "Canonical URL"
            canonical_url,
        ]

        # Add parent row ONLY if no variants
        if not product.variants:
            rows.append(parent_row)

        if product.variants:
            for variant in product.variants:
                # Attributes -> up to 2 options
                attrs = list(variant.attributes or [])
                attr1_name = attrs[0].name if len(attrs) >= 1 else ""
                attr1_val = attrs[0].value if len(attrs) >= 1 else ""
                attr2_name = attrs[1].name if len(attrs) >= 2 else ""
                attr2_val = attrs[1].value if len(attrs) >= 2 else ""

                # Price / RRP / stock / dimensions / shipping
                selling_price = variant.price or product.price
                rrp = variant.rrp_price or product.rrp_price
                cost_price = variant.cost_price or product.cost_price
                stock_qty = (
                    variant.stock if variant.stock is not None else product.stock
                )

                weight = (
                    variant.weight
                    if variant.weight not in (None, 0)
                    else product.weight
                )
                length = (
                    variant.length
                    if variant.length not in (None, 0)
                    else product.length
                )
                width = (
                    variant.width if variant.width not in (None, 0) else product.width
                )
                height = (
                    variant.height
                    if variant.height not in (None, 0)
                    else product.height
                )

                handling_time = (
                    variant.handling_time_days
                    if variant.handling_time_days is not None
                    else product.handling_time_days
                )

                ships_from_location = (
                    variant.ships_from_location
                    if variant.ships_from_location
                    else product.ships_from_location or ""
                )

                # Images & video (variant overrides)
                img_urls = _collect_image_urls(product, variant)
                video_url = _collect_video_url(product, variant) or product_level_video

                main_image = img_urls[0] if len(img_urls) >= 1 else ""
                lifestyle_image = img_urls[1] if len(img_urls) >= 2 else ""

                extra_imgs = img_urls[2:12]  # up to 10 images total
                extra_imgs += [""] * (10 - len(extra_imgs))  # pad to 10

                # Build row in EXACT header order
                row = [
                    # 1–3: category info
                    category_code or "",  # "category code"
                    category_breadcrumb or "",  # "category"
                    leaf_product_type or "",  # "product type"
                    # 4. "SKU"
                    variant.sku or product.sku or "",
                    # 5. "Product Code"
                    variant.ean or "",
                    # 6. "Title"
                    product.title or "",
                    # 7. "Long Description"
                    product.description or "",
                    # 8. "Brand_id"
                    product.brand.name if product.brand else "",
                    # 9. "Condition"
                    product.product_condition or "",
                    # 10. "Product Status"
                    product.status or "",
                    # 11. "Supplier Name"
                    supplier_name,
                    # 12. "Product Tag"
                    product_tag,
                    # 13. "Parent SKU / Variation Group Code"
                    product.sku or "",
                    # 14. "Variant Option 1"
                    attr1_name,
                    # 15. "Variant Values 1"
                    attr1_val,
                    # 16. "Variant Option 2"
                    attr2_name,
                    # 17. "Variant Values 2"
                    attr2_val,
                    # 18. "Bundle / Combo Indicator"
                    variant.bundle_group_code or "",
                    # 19. "Selling Price"
                    _dec_or_empty(selling_price),
                    # 20. "RRP"
                    _dec_or_empty(rrp),
                    # 21. "Cost Price"
                    _dec_or_empty(cost_price),
                    # 22. "Stock Quantity"
                    str(stock_qty or ""),
                    # 23. "Package Weight (kg)"
                    _dec_or_empty(weight),
                    # 24. "Package Length(Cms)"
                    _dec_or_empty(length),
                    # 25. "Package Width(Cms)"
                    _dec_or_empty(width),
                    # 26. "Package Height(Cms)"
                    _dec_or_empty(height),
                    # 27. "Harmonized Code (HS Code)"
                    hs_code,
                    # 28. "Shipping Template"
                    shipping_template,
                    # 29. "Handling Time (days)"
                    str(handling_time or "") if handling_time is not None else "",
                    # 30. "Fast Dispatch"
                    product.fast_dispatch if product.fast_dispatch is not None else "",
                    # 31. "Free Shipping"
                    product.free_shipping if product.free_shipping is not None else "",
                    # 32. "Ships From Location"
                    ships_from_location,
                    # 33. "Product Video Link"
                    video_url,
                    # 34. "Main Image URL"
                    main_image,
                    # 35. "Lifestyle Image URL"
                    lifestyle_image,
                    # 36–45. "Image 1 URL"..."Image 10 URL"
                    *extra_imgs,
                    # 46. "Precautionary Note"
                    precautionary_note,
                    # 47. "Care Instructions (If Applicable)"
                    care_instructions,
                    # 48. "Warranty / Guarantee (If Applicable)"
                    warranty,
                    # 49. "SEO Keywords"
                    seo_keywords,
                    # 50. "Page Title"
                    page_title,
                    # 51. "Meta Description"
                    meta_description,
                    # 52. "URL Handles"
                    url_handle,
                    # 53. "Canonical URL"
                    canonical_url,
                ]

                rows.append(row)

        processed_count += 1
        if task_id and processed_count % 50 == 0:
            await db.execute(
                update(BackgroundTask)
                .where(BackgroundTask.task_id == task_id)
                .values(task_info={"total_rows": total_products, "processed_rows": processed_count})
            )
            await db.commit()

    if task_id:
        await db.execute(
            update(BackgroundTask)
            .where(BackgroundTask.task_id == task_id)
            .values(task_info={"total_rows": total_products, "processed_rows": total_products})
        )
        await db.commit()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    print("BEFORE HEADERS")
    CATEGORY_HEADERS = ["Category code", "category"]
    new_headers = EXPECTED_HEADERS.copy()
    new_headers[0] = "Product Type"
    new_headers.pop(1)
    print("NEW HEADERS", new_headers)
    export_headers = CATEGORY_HEADERS + new_headers
    print("EXPORT HEADERS", export_headers)
    # -------- COLUMN FILTERING --------
    if columns and len(columns) > 0:
        requested = [c.strip() for c in columns]
        REQUIRED_COLUMNS = [
            "Category code",
            "category",
            "Product Type",
            "Product Code",
            "Title",
            "SKU",
            "Brand Name",
            "Parent SKU / Variation Group Code",
        ]
        # merge required + requested (preserve order, avoid duplicates)
        final_columns = []
        for c in REQUIRED_COLUMNS + requested:
            if c in export_headers and c not in final_columns:
                final_columns.append(c)

        # build indices based on final_columns
        indices = [i for i, h in enumerate(export_headers) if h in final_columns]
        print("INDICES", indices)
        export_headers = [export_headers[i] for i in indices]
        rows = [[row[i] for i in indices] for row in rows]
        print("ROWS", rows)
    print("EXPORT HEADERS", export_headers)
    return export_headers, rows, timestamp


def write_file_for_response(headers, rows, timestamp, download_flag):
    headers = headers + ["action"]
    rows = [row + [""] for row in rows]
    if download_flag == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        output.seek(0)
        return output, f"products_export_{timestamp}.csv", "text/csv"

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(headers)

    bold = Font(bold=True)
    for c in ws[1]:
        c.font = bold
    for r in rows:
        ws.append(r)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return (
        stream,
        f"products_export_{timestamp}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def build_file_bytes(headers, rows, timestamp, download_flag):
    """
    Builds CSV/XLSX fully in memory.
    Used ONLY by Celery.
    """
    headers = headers + ["action"]
    rows = [row + [""] for row in rows]

    if download_flag == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        writer.writerows(rows)

        return (
            buffer.getvalue().encode("utf-8"),
            f"products_export_{timestamp}.csv",
            "text/csv",
        )

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)

    return (
        buffer.getvalue(),
        f"products_export_{timestamp}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
