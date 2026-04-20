from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from deps import get_db
from models.brand import Brand
from schemas.brand import BrandCreate, BrandUpdate, BrandOut
from fastapi import status
from fastapi import Query
import pandas as pd
from fastapi import UploadFile, File
from schemas.brand import BrandStatusUpdate, BrandOutStatus
from sqlalchemy import func
from typing import Annotated, Optional
from models.product import Product
from sqlalchemy import case
import io
import csv
from fastapi.responses import StreamingResponse
from apis.v1.import_export import make_slug
import utils.constants as constants

from utils.activity_logger import log_activity
from typing import Optional, List, Generic, TypeVar, Literal
from openpyxl import Workbook
from openpyxl.styles import Font
from apis.v1.utils import make_slug
import utils.constants as constants

from pydantic import BaseModel
from typing import List, Generic, TypeVar
from utils.activity_logger import log_activity
from service.brand_admin import (
    create_brand as _create_brand,
    list_brands as _list_brands,
    _get_brand,
    _update_brand,
    _delete_brand,
)


router = APIRouter()

T = TypeVar("T")


class PaginatedResponse(BaseModel, Generic[T]):
    page: int
    limit: int
    total: int
    pages: int
    data: List[T]


# Create a brand
@router.post(
    "/create-brand",
    response_model=BrandOut,
    responses={
        400: {"description": "Brand with the same name already exists."},
    },
)
async def create_brand(
    brand: BrandCreate, db: Annotated[AsyncSession, Depends(get_db)]
):
    new_brand = await _create_brand(db, brand)
    return new_brand


@router.get("/list-brands", response_model=PaginatedResponse[BrandOut])
async def get_brands(
    db: Annotated[AsyncSession, Depends(get_db)],
    page: Annotated[int, Query(ge=1)] = 1,
    size: Annotated[int, Query(ge=1, le=1000)] = 10,
    status: Annotated[str, Query(pattern="^(active|inactive|all)$")] = "active",
    sort_by: Annotated[str, Query(pattern="^(name|created_at)$")] = "name",
    sort_dir: Annotated[str, Query(pattern="^(asc|desc)$")] = "asc",
    download: Annotated[str, Query(pattern="^(yes|no)$")] = "no",
    q: Annotated[Optional[str], Query(description="Search brand by name")] = None,
    only_with_products: Annotated[
        bool, Query(description="Filter brands with at least one product")
    ] = False,
):
    total, rows = await _list_brands(
        db, page, size, status, q, sort_by, sort_dir, only_with_products, download
    )
    return PaginatedResponse(
        total=total,
        page=page,
        limit=size,
        pages=(total + size - 1) // size,
        data=rows,
    )


# Get a brand by ID
@router.get(
    "/get-brand/{brand_id}",
    response_model=BrandOut,
    responses={
        404: {"description": "Brand not found."},
    },
)
async def get_brand(brand_id: str, db: Annotated[AsyncSession, Depends(get_db)]):
    brand = await _get_brand(db, brand_id)
    return brand


# Update a brand
@router.put(
    "/update-brand/{brand_id}",
    response_model=BrandOut,
    responses={
        400: {"description": "Brand name already exists."},
        404: {"description": "Brand not found."},
    },
)
async def update_brand(
    brand_id: str, brand_data: BrandUpdate, db: Annotated[AsyncSession, Depends(get_db)]
):
    brand = await _update_brand(db, brand_id, brand_data)
    return brand


# Delete a brand
@router.delete(
    "/delete-brand/{brand_id}",
    responses={
        400: {"description": "Cannot delete brand with associated products."},
        404: {"description": "Brand not found."},
    },
)
async def delete_brand(brand_id: str, db: Annotated[AsyncSession, Depends(get_db)]):
    await _delete_brand(db, brand_id)
    return {"detail": "Brand deleted"}


@router.post(
    "/upload-brands",
    status_code=status.HTTP_201_CREATED,
    responses={
        400: {
            "description": "Unsupported file format, error reading file, or missing columns."
        },
    },
)
async def upload_brands(
    file: Annotated[UploadFile, File(...)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Upload a CSV or Excel file with columns:
    - Brand name (column name flexible, e.g. 'brand', 'Brand Name', etc.)
    - logo_url (optional)
    """

    # --- Step 1: Load data ---
    try:
        if file.filename.endswith(".csv"):
            df = pd.read_csv(file.file)
        elif file.filename.endswith((".xlsx", ".xls", ".xlsb")):
            df = pd.read_excel(
                file.file,
                engine="pyxlsb" if file.filename.endswith(".xlsb") else None,
            )
        else:
            raise HTTPException(
                status_code=400, detail="Unsupported file format. Use CSV or Excel."
            )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading file: {e}")

    # --- Step 2: Normalize column names ---
    df.columns = df.columns.str.strip().str.lower()

    # Try to find the brand name column dynamically
    possible_brand_cols = ["brand", "brand name", "brand_name", "name"]
    brand_col = next((col for col in df.columns if col in possible_brand_cols), None)

    if not brand_col:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required brand name column. Expected one of: {possible_brand_cols}",
        )

    # --- Step 3: Clean data ---
    df[brand_col] = df[brand_col].astype(str).str.strip()
    df = df.dropna(subset=[brand_col])
    df = df.drop_duplicates(subset=[brand_col])

    # --- Step 4: Fetch existing brand names ---
    existing_result = await db.execute(select(Brand.name))
    existing_names = {row[0].lower() for row in existing_result.all()}

    # --- Step 5: Prepare new brands ---
    new_brands = []
    for _, row in df.iterrows():
        name = row[brand_col].strip()
        if name.lower() in existing_names:
            continue

        logo_url = None
        # Detect logo column if present
        for col in df.columns:
            if col in ["logo_url", "logo", "logo link", "image", "logo url"]:
                logo_url = row.get(col)
                break

        new_brands.append(
            Brand(
                name=name,
                logo_url=logo_url,
                is_active=True,
            )
        )

    if not new_brands:
        return {"message": "No new brands to add. All already exist."}

    # --- Step 6: Commit to DB ---
    db.add_all(new_brands)
    await db.commit()

    return {
        "message": f"Successfully added {len(new_brands)} new brands.",
        "total_uploaded": len(df),
        "total_added": len(new_brands),
    }


@router.patch(
    "/status/{brand_id}/",
    response_model=BrandOutStatus,
    responses={
        404: {"description": "Brand not found."},
    },
)
async def update_brand_status(
    brand_id: str, data: BrandStatusUpdate, db: Annotated[AsyncSession, Depends(get_db)]
):
    # Fetch brand
    result = await db.execute(select(Brand).where(Brand.id == brand_id))
    brand = result.scalar_one_or_none()

    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")

    # Update status
    brand.is_active = data.is_active
    brand.updated_at = func.now()

    await db.commit()
    await db.refresh(brand)

    return brand


@router.get("/search")
async def search_brands(
    db: Annotated[AsyncSession, Depends(get_db)],
    q: Annotated[Optional[str], Query(description="Search brand by name")] = None,
    page: Annotated[int, Query(ge=1)] = 1,
    limit: Annotated[int, Query(ge=1, le=100)] = 10,
):
    offset = (page - 1) * limit

    # ---- BASE QUERY (Only Active Brands) ----
    query = select(
        Brand.id,
        Brand.name,
        Brand.logo_url,
        Brand.is_active,
        Brand.created_at,
        Brand.updated_at,
    ).where(Brand.is_active == True)

    # ---- APPLY SEARCH ----
    if q:
        search_term = f"%{q.strip()}%"
        query = query.where(Brand.name.ilike(search_term))

    # ---- COUNT ----
    total_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(total_query)).scalar_one()

    # ---- PAGINATION ----
    result = await db.execute(
        query.order_by(Brand.name.asc()).limit(limit).offset(offset)
    )

    rows = result.mappings().all()

    return {
        "query": q,
        "page": page,
        "limit": limit,
        "total": total,
        "pages": (total + limit - 1) // limit,
        "count": len(rows),
        "results": rows,
    }


# IMPORT EXPORT


@router.get("/download-template")
async def download_brand_template(
    file_type: Annotated[Literal["csv", "excel"], Query()] = "excel",
):
    headers = ["Brand Name", "logo url", "image url", "Action"]
    filename = "brand_upload_template"

    if file_type == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )

    # Excel
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Brands"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


@router.get("/export-brands", response_class=StreamingResponse)
async def export_brands(
    db: Annotated[AsyncSession, Depends(get_db)],
    file_type: Annotated[Literal["csv", "excel"], Query()] = "excel",
):
    # Fetch all brands
    result = await db.execute(
        select(
            Brand.id,
            Brand.name,
            Brand.slug,
            Brand.logo_url,
            Brand.image_url,
            Brand.is_active,
            func.count(Product.id).label("product_count"),
        )
        .outerjoin(Product, Product.brand_id == Brand.id)
        .group_by(Brand.id)
    )

    brands = result.mappings().all()

    headers = [
        "Brand ID",
        "Brand Name",
        "Slug",
        "Logo URL",
        "Image URL",
        "Status",
        "Product Count",
        "Action",
    ]

    from datetime import datetime

    filename = f"brands_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if file_type == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for b in brands:
            writer.writerow(
                [
                    b["id"],
                    b["name"],
                    b["slug"],
                    b["logo_url"],
                    b["image_url"],
                    "Active" if b["is_active"] else "Inactive",
                    b["product_count"],
                    "",
                ]
            )
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )

    # Excel
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Brands"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for b in brands:
        ws.append(
            [
                b["id"],
                b["name"],
                b["slug"],
                b["logo_url"],
                b["image_url"],
                "Active" if b["is_active"] else "Inactive",
                b["product_count"],
                "",
            ]
        )

    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


def check_file_type(file: UploadFile = File(...)):
    if file.filename.endswith(".csv"):
        df = pd.read_csv(file.file)
    elif file.filename.endswith((".xlsx", ".xls", ".xlsb")):
        df = pd.read_excel(
            file.file,
            engine="pyxlsb" if file.filename.endswith(".xlsb") else None,
        )
    else:
        raise HTTPException(
            status_code=400,
            detail=constants.messages.get(
                "unsupported_file_format", "Unsupported file format. Use CSV or Excel."
            ),
        )
    return df


@router.post(
    "/upload-brands",
    status_code=status.HTTP_201_CREATED,
    responses={
        400: {
            "description": "Unsupported file format, error reading file, or missing columns."
        },
    },
)
async def upload_brands(
    file: Annotated[UploadFile, File(...)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Upload a CSV or Excel file with columns:
    - Brand name (column name flexible, e.g. 'brand', 'Brand Name', etc.)
    - logo_url (optional)
    """

    # --- Step 1: Load data ---
    try:
        df = check_file_type(file)
    except Exception as e:
        print("################", e)
        raise HTTPException(status_code=400, detail=f"Error reading file: {e}")

    # --- Step 2: Normalize column names ---
    df.columns = df.columns.str.strip().str.lower()
    # Replace NaN/NaT with None for DB compatibility
    df = df.where(pd.notnull(df), None)

    # Validate columns
    MANDATORY_COLUMNS = {"brand name", "logo url", "image url"}
    missing_cols = [col for col in MANDATORY_COLUMNS if col not in df.columns]
    if missing_cols:
        print("################", missing_cols)
        raise HTTPException(
            status_code=400,
            detail=constants.messages.get(
                "missing_mandatory_columns",
                "Missing mandatory columns: {missing_cols}. Please use the standard template.",
            ).format(missing_cols=missing_cols),
        )

    # --- Step 3: Clean data ---
    brand_col = "brand name"
    action_col = "action"

    # Ensure brand name is string
    df[brand_col] = df[brand_col].apply(
        lambda x: str(x).strip() if x is not None else None
    )

    # Ensure action column exists (optional)
    if action_col in df.columns:
        df[action_col] = df[action_col].apply(
            lambda x: str(x).strip().lower() if x is not None else ""
        )
    else:
        df[action_col] = ""

    # Drop rows where brand name is None
    df = df.dropna(subset=[brand_col])
    # Filter out any "nan" or "none" strings
    df = df[~df[brand_col].str.lower().isin(["nan", "none", ""])]
    df = df.drop_duplicates(subset=[brand_col])

    # --- Step 4: Fetch existing brands ---
    existing_result = await db.execute(select(Brand))
    existing_brands_list = existing_result.scalars().all()
    existing_brands_map = {b.name.lower(): b for b in existing_brands_list}

    # --- Step 5: Process rows ---
    new_brands = []
    deleted_count = 0
    updated_count = 0
    skipped_count = 0
    errors = []

    for _, row in df.iterrows():
        name = row[brand_col].strip()
        action = row[action_col]

        brand_obj = existing_brands_map.get(name.lower())

        if action == "delete":
            if not brand_obj:
                errors.append(f"Brand '{name}' not found for deletion.")
                continue

            # Check for products
            product_count = await db.scalar(
                select(func.count(Product.id)).where(Product.brand_id == brand_obj.id)
            )
            if product_count > 0:
                errors.append(
                    f"Cannot delete brand '{name}' with {product_count} associated products."
                )
                continue

            await db.delete(brand_obj)
            deleted_count += 1

        else:  # Create or Update
            logo_url = row.get("logo url")
            image_url = row.get("image url")

            # Clean URLs
            if logo_url and str(logo_url).lower() in ["nan", "none", ""]:
                logo_url = None
            if image_url and str(image_url).lower() in ["nan", "none", ""]:
                image_url = None

            if brand_obj:
                # Compare for updates
                has_changes = False
                if brand_obj.logo_url != logo_url:
                    brand_obj.logo_url = logo_url
                    has_changes = True
                if brand_obj.image_url != image_url:
                    brand_obj.image_url = image_url
                    has_changes = True

                if has_changes:
                    updated_count += 1
                else:
                    skipped_count += 1
                continue

            new_brands.append(
                Brand(
                    name=name,
                    slug=make_slug(name),
                    logo_url=logo_url,
                    image_url=image_url,
                    is_active=True,
                )
            )

    if new_brands:
        db.add_all(new_brands)

    await db.commit()

    return {
        "message": "Upload processing complete.",
        "created": len(new_brands),
        "updated": updated_count,
        "deleted": deleted_count,
        "skipped": skipped_count,
        "errors": errors,
        "total_rows": len(df),
    }
