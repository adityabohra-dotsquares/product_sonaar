# apis/v1/product_import_export.py
import io
import os
import asyncio
import pandas as pd
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
import re
import unicodedata
from models.product import (
    Product,
    ProductVariant,
    Attribute,
    ProductImage,
    ProductVariantImage,
)
from models.review import Review
from models.seo import ProductSEO
from deps import get_db
from utils.admin_auth import require_catalog_supervisor
from utils.activity_logger import log_activity
from service.review_stats import update_review_stats
from fastapi.responses import StreamingResponse
import csv
from datetime import datetime
from fastapi import status
from sqlalchemy.orm import selectinload

from models.category import Category
from models.brand import Brand
from typing import List, Dict, Any, Annotated
from io import StringIO, BytesIO
from openpyxl import load_workbook
from decimal import Decimal
from typing import Literal
from fastapi import APIRouter, Query, Depends
from sqlalchemy import select, delete, update
from typing import Tuple
from openpyxl import Workbook
from typing import Optional, List, Literal
from openpyxl.styles import Font
from schemas.import_export import (
    ProductImportRow,
    format_pydantic_error,
    BulkTemplateRequest,
)
from .utils import (
    safe_get,
    make_slug,
    validate_brand_and_category,
    clean_str_with_strip,
    clean_str,
    to_int,
    to_decimal,
    to_bool,
    redis_client,
)
from utils.gcp_bucket import upload_file_to_gcs
from utils.image_handler import download_and_upload_image
from models.b_tasks import BackgroundTask


from celery_worker.celery_app import app
import logging
import asyncio
import anyio

logger = logging.getLogger(__name__)


router = APIRouter()


async def _read_file_rows(file: str | UploadFile) -> List[Dict[str, Any]]:
    # Read file contents
    if isinstance(file, str):
        if not await anyio.Path(file).exists():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="File does not exist."
            )
        contents = await anyio.Path(file).read_bytes()
        filename = os.path.basename(file)
    else:
        contents = await file.read()
        filename = file.filename if file.filename else "temp_file"

    # Get file extension
    ext = filename.split(".")[-1].lower()  # Extract file extension

    # CSV
    if ext == "csv":
        text = contents.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        reader.fieldnames = [
            (h.strip().lower() if h is not None else "") for h in reader.fieldnames
        ]
        rows: List[Dict[str, Any]] = []
        for row in reader:
            normalized_row = {
                (k.strip().lower() if k else ""): v for k, v in row.items()
            }
            # Skip rows where all values are empty
            if all(
                v is None or (isinstance(v, str) and v.strip() == "")
                for v in normalized_row.values()
            ):
                continue
            rows.append(normalized_row)
        return rows

    # Excel (.xls or .xlsx)
    if ext in ("xls", "xlsx"):
        wb = load_workbook(BytesIO(contents), read_only=True)

        sheet_name = None
        # Search for "products template" sheet, else fallback to the first sheet
        for name in wb.sheetnames:
            if name.strip().lower() == "products template":
                sheet_name = name
                break

        # Use the selected sheet, or fallback to the active sheet
        ws = wb[sheet_name] if sheet_name else wb.active

        rows_iter = ws.iter_rows(values_only=True)

        try:
            headers = [
                str(h).strip().lower() if h is not None else "" for h in next(rows_iter)
            ]
        except StopIteration:
            return []

        rows: List[Dict[str, Any]] = []
        for r in rows_iter:
            # Skip rows where all values are empty or None
            if r is None or all(
                (cell is None or str(cell).strip() == "") for cell in r
            ):
                continue

            row_dict: Dict[str, Any] = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                value = r[i] if i < len(r) else None
                row_dict[header] = value

            # Skip rows where all values are empty
            if all(
                v is None or (isinstance(v, str) and v.strip() == "")
                for v in row_dict.values()
            ):
                continue

            rows.append(row_dict)

        return rows

    # If the file format is unsupported
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Unsupported file type. Please upload .csv, .xls, or .xlsx.",
    )


BATCH_SIZE = 300


@router.post(
    "/import-products",
    status_code=202,
    responses={
        400: {"description": "Invalid file format or row limit exceeded."},
    },
)
async def import_products(
    file: Annotated[UploadFile, File(...)],
    db: Annotated[AsyncSession, Depends(get_db)],
    user: Annotated[dict, Depends(require_catalog_supervisor)],
):
    import uuid

    job_id = f"import_{uuid.uuid4()}"
    blob_name = f"imports/{job_id}.xlsx"

    # read file bytes
    file_bytes = await file.read()

    # Check row count limit
    filename = file.filename
    ext = filename.split(".")[-1].lower() if filename else ""

    MAX_ROWS = 10000
    LIMIT = MAX_ROWS + 1  # 1 header row + 10,000 data rows

    if ext in ("xls", "xlsx"):
        try:
            wb = load_workbook(BytesIO(file_bytes), read_only=True)
            sheet_name = None
            for name in wb.sheetnames:
                if name.strip().lower() == "products template":
                    sheet_name = name
                    break
            ws = wb[sheet_name] if sheet_name else wb.active

            count = 0
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    count += 1
            if count > LIMIT:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"File exceeds the maximum limit of {MAX_ROWS} rows.",
                )
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Failed to read Excel file: {str(e)}",
            )

    elif ext == "csv":
        try:
            text = file_bytes.decode("utf-8-sig")
            reader = csv.reader(StringIO(text))
            count = 0
            for row in reader:
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    count += 1
                    if count > LIMIT:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f"File exceeds the maximum limit of {MAX_ROWS} rows.",
                        )
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Failed to read CSV file: {str(e)}",
            )

    # upload to GCS
    from utils.gcp_bucket import upload_file_to_gcs  # adjust to your path

    file_url = upload_file_to_gcs(
        file_bytes=file_bytes,
        content_type=file.content_type,
        blob_name=blob_name,
    )

    print("Importing products from GCS:", file_url)
    redis_client.hset(
        job_id,
        mapping={
            "status": "processing",
            "processed_rows": 0,
            "created_products": 0,
            "updated_products": 0,
            "created_variants": 0,
            "updated_variants": 0,
            "errors": "[]",
        },
    )
    bg_task = BackgroundTask(
        task_id=job_id,
        task_type="IMPORT_PRODUCTS",
        status="PENDING",
        added_by="system",
        file_url=file_url,
    )
    db.add(bg_task)
    await db.commit()
    from celery_worker.celery_app import import_products_task

    res = import_products_task.delay(file_url, job_id)

    return {"job_id": job_id, "status": "queued"}


@router.get(
    "/import-products/status/{job_id}",
    responses={
        404: {"description": "Invalid job id."},
    },
)
async def import_status(job_id: str):
    data = redis_client.hgetall(job_id)

    if not data:
        raise HTTPException(status_code=404, detail="Invalid job id")

    decoded = {k.decode(): v.decode() for k, v in data.items()}

    if "processed_rows" in decoded and "total_rows" in decoded:
        decoded["progress"] = round(
            (int(decoded["processed_rows"]) / int(decoded["total_rows"])) * 100, 2
        )

    return decoded


# --------------------------------------------------
# PAUSE / RESUME / ABORT
# --------------------------------------------------


@router.post(
    "/import-products/{job_id}/pause",
    status_code=200,
    responses={
        404: {"description": "Invalid job id."},
    },
)
async def pause_import(job_id: str, db: Annotated[AsyncSession, Depends(get_db)]):
    """Pause a running import job. The worker stops before the next batch."""
    data = redis_client.hgetall(job_id)
    if not data:
        raise HTTPException(status_code=404, detail="Invalid job id")

    redis_client.set(f"{job_id}:ctrl", "pause")

    await db.execute(
        update(BackgroundTask)
        .where(BackgroundTask.task_id == job_id)
        .values(status="PAUSED")
    )
    await db.commit()

    return {"job_id": job_id, "status": "PAUSED"}


@router.post(
    "/import-products/{job_id}/resume",
    status_code=200,
    responses={
        404: {"description": "Invalid job id."},
    },
)
async def resume_import(job_id: str, db: Annotated[AsyncSession, Depends(get_db)]):
    """Resume a paused import job."""
    data = redis_client.hgetall(job_id)
    if not data:
        raise HTTPException(status_code=404, detail="Invalid job id")

    redis_client.set(f"{job_id}:ctrl", "running")

    await db.execute(
        update(BackgroundTask)
        .where(BackgroundTask.task_id == job_id)
        .values(status="RUNNING")
    )
    await db.commit()

    return {"job_id": job_id, "status": "RUNNING"}


@router.post(
    "/import-products/{job_id}/abort",
    status_code=200,
    responses={
        404: {"description": "Invalid job id."},
    },
)
async def abort_import(job_id: str, db: Annotated[AsyncSession, Depends(get_db)]):
    """
    Abort a running or paused import job.
    Sets the Redis control flag to 'abort' so the worker exits cleanly,
    and revokes the Celery task (hard kill) as a safety net.
    """
    result = await db.execute(
        select(BackgroundTask).where(BackgroundTask.task_id == job_id)
    )
    task = result.scalar_one_or_none()
    print("task", task)
    if not task:
        raise HTTPException(status_code=404, detail="Invalid job id")

    # 1. Signal the worker to stop
    redis_client.set(f"{job_id}:ctrl", "abort")

    # 2. Hard-terminate the Celery task (safety net if worker is blocked on I/O)
    from celery_worker.celery_app import app as celery_app

    celery_app.control.revoke(job_id, terminate=True, signal="SIGTERM")

    # 3. Update DB (worker will also update, but do it here as well in case
    #    the worker has already exited or was never started)
    await db.execute(
        update(BackgroundTask)
        .where(BackgroundTask.task_id == job_id)
        .values(status="ABORTED")
    )
    await db.commit()

    return {"job_id": job_id, "status": "ABORTED"}


@router.post(
    "/import-products-ex",
    status_code=status.HTTP_201_CREATED,
    responses={
        400: {
            "description": "File is empty, unparsable, or contains validation errors."
        },
    },
)
async def import_products(
    file: Annotated[UploadFile, File(...)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    rows = await _read_file_rows(file)
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File is empty or could not be parsed.",
        )
    errors: List[str] = []
    created_products = 0
    updated_products = 0
    created_variants = 0
    updated_variants = 0
    # Validate rows with Pydantic
    for idx, row in enumerate(rows):
        try:
            validated = ProductImportRow(**row)
        except Exception as e:
            simple_errors = format_pydantic_error(e)
            for m in simple_errors:
                errors.append(f"Row {idx + 2}: {m}")
            continue
    if errors:
        return {
            "created_products": created_products,
            "updated_products": updated_products,
            "created_variants": created_variants,
            "updated_variants": updated_variants,
            "errors": errors,
        }
    is_valid, validation_errors, brand_map, category_map = (
        await validate_brand_and_category(db, rows, Brand, Category)
    )
    if not is_valid:
        errors.extend(validation_errors)
        return {
            "created_products": created_products,
            "updated_products": updated_products,
            "created_variants": created_variants,
            "updated_variants": updated_variants,
            "errors": errors,
        }
    # Group rows by variation group code or SKU
    await db.commit()
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for idx, row in enumerate(rows):
        sku_raw = row.get("sku")
        parent_code_raw = row.get("parent sku / variation group code")
        if parent_code_raw:
            group_key = clean_str_with_strip(parent_code_raw)
        else:
            if not sku_raw:
                errors.append(
                    f"Row {idx + 2}: No SKU or parent code provided. Skipping."
                )
                continue
            group_key = clean_str_with_strip(sku_raw)
        grouped.setdefault(group_key, []).append(row)
    # Image columns
    image_columns = [
        "main image url",
        "lifestyle image url",
        "image 1 url",
        "image 2 url",
        "image 3 url",
        "image 4 url",
        "image 5 url",
        "image 6 url",
        "image 7 url",
        "image 8 url",
        "image 9 url",
        "image 10 url",
    ]
    # Essential fields for creation
    essential_product_fields = ["title", "selling price"]
    essential_variant_fields = ["title", "selling price"]
    async with db.begin():
        for product_sku, sku_rows in grouped.items():
            base_row = sku_rows[0]
            parent_code_raw = base_row.get("parent sku / variation group code")
            variation_group_code = (
                clean_str_with_strip(parent_code_raw) if parent_code_raw else None
            )
            product_code_raw = base_row.get("product code")
            product_code = (
                clean_str_with_strip(product_code_raw)
                if product_code_raw is not None
                else None
            )
            # Check product_code uniqueness if provided
            if product_code:
                result_code = await db.execute(
                    select(Product).where(Product.product_code == product_code)
                )
                existing_by_code = result_code.scalar_one_or_none()
                if existing_by_code:
                    errors.append(
                        f"Product code '{product_code}' already exists. Skipping group."
                    )
                    continue
            # Check existing product by SKU
            result_sku = await db.execute(
                select(Product).where(Product.sku == product_sku)
            )
            existing_product = result_sku.scalar_one_or_none()
            current_product = None
            updated = False
            if existing_product:
                # Update existing product with provided fields only
                # title and slug
                title_raw = base_row.get("title")
                if title_raw is not None:
                    new_title = clean_str_with_strip(title_raw)
                    if existing_product.title != new_title:
                        existing_product.title = new_title
                        existing_product.slug = (
                            make_slug(clean_str(new_title))
                            if new_title
                            else existing_product.slug
                        )
                        updated = True
                # description
                desc_raw = base_row.get("long description")
                if desc_raw is not None:
                    new_desc = clean_str_with_strip(desc_raw)
                    if existing_product.description != new_desc:
                        existing_product.description = new_desc
                        updated = True
                # price
                price_raw = base_row.get("selling price")
                if price_raw is not None:
                    new_price = to_decimal(price_raw)
                    if existing_product.price != new_price:
                        existing_product.price = new_price
                        updated = True
                # product_type
                type_raw = base_row.get("product_type")
                if type_raw is not None:
                    new_type = clean_str_with_strip(type_raw)
                    if existing_product.product_type != new_type:
                        existing_product.product_type = new_type
                        updated = True
                # cost_price
                cost_raw = base_row.get("cost price")
                if cost_raw is not None:
                    new_cost = to_decimal(cost_raw)
                    if existing_product.cost_price != new_cost:
                        existing_product.cost_price = new_cost
                        updated = True
                # rrp_price
                rrp_raw = base_row.get("rrp")
                if rrp_raw is not None:
                    new_rrp = to_decimal(rrp_raw)
                    if existing_product.rrp_price != new_rrp:
                        existing_product.rrp_price = new_rrp
                        updated = True
                # status
                status_raw = base_row.get("product status")
                if status_raw is not None:
                    new_status = clean_str(status_raw)
                    if existing_product.status != new_status:
                        existing_product.status = new_status
                        updated = True
                # weight
                weight_raw = base_row.get("package weight (kg)")
                if weight_raw is not None:
                    new_weight = to_decimal(weight_raw)
                    if existing_product.weight != new_weight:
                        existing_product.weight = new_weight
                        updated = True
                # length
                length_raw = base_row.get("package length(cms)")
                if length_raw is not None:
                    new_length = to_decimal(length_raw)
                    if existing_product.length != new_length:
                        existing_product.length = new_length
                        updated = True
                # width
                width_raw = base_row.get("package width(cms)")
                if width_raw is not None:
                    new_width = to_decimal(width_raw)
                    if existing_product.width != new_width:
                        existing_product.width = new_width
                        updated = True
                # height
                height_raw = base_row.get("package height(cms)")
                if height_raw is not None:
                    new_height = to_decimal(height_raw)
                    if existing_product.height != new_height:
                        existing_product.height = new_height
                        updated = True
                # bundle_group_code
                bundle_raw = base_row.get("bundle / combo indicator")
                if bundle_raw is not None:
                    new_bundle = clean_str_with_strip(bundle_raw)
                    if existing_product.bundle_group_code != new_bundle:
                        existing_product.bundle_group_code = new_bundle
                        updated = True
                # supplier
                supplier_raw = base_row.get("supplier name")
                if supplier_raw is not None:
                    new_supplier = clean_str_with_strip(supplier_raw)
                    if existing_product.supplier != new_supplier:
                        existing_product.supplier = new_supplier
                        updated = True
                # country_of_origin
                country_raw = base_row.get("country of origin")
                if country_raw is not None:
                    new_country = clean_str_with_strip(country_raw)
                    if existing_product.country_of_origin != new_country:
                        existing_product.country_of_origin = new_country
                        updated = True
                # product_id_type
                ptype_raw = base_row.get("product id type")
                if ptype_raw is not None:
                    new_ptype = clean_str_with_strip(ptype_raw)
                    if existing_product.product_id_type != new_ptype:
                        existing_product.product_id_type = new_ptype
                        updated = True
                # hs_code
                hs_raw = base_row.get("harmonized code (hs code)")
                if hs_raw is not None:
                    new_hs = clean_str_with_strip(hs_raw)
                    if existing_product.hs_code != new_hs:
                        existing_product.hs_code = new_hs
                        updated = True
                # category_id
                cat_code_raw = base_row.get("category code")
                if cat_code_raw is not None:
                    new_cat_id = safe_get(category_map, clean_str(cat_code_raw))
                    if new_cat_id and existing_product.category_id != new_cat_id:
                        existing_product.category_id = new_cat_id
                        updated = True
                # brand_id
                brand_raw = base_row.get("brand name")
                if brand_raw is not None:
                    new_brand_id = safe_get(brand_map, clean_str(brand_raw))
                    if new_brand_id and existing_product.brand_id != new_brand_id:
                        existing_product.brand_id = new_brand_id
                        updated = True
                # stock
                stock_raw = base_row.get("stock quantity")
                if stock_raw is not None:
                    new_stock = to_int(stock_raw)
                    if existing_product.stock != new_stock:
                        existing_product.stock = new_stock
                        updated = True
                # is_battery_required
                battery_raw = base_row.get("is battery required?")
                if battery_raw is not None:
                    new_battery = to_bool(battery_raw)
                    if existing_product.is_battery_required != new_battery:
                        existing_product.is_battery_required = new_battery
                        updated = True
                # shipping_template
                ship_temp_raw = base_row.get("shipping template")
                if ship_temp_raw is not None:
                    new_ship_temp = clean_str_with_strip(ship_temp_raw)
                    if existing_product.shipping_template != new_ship_temp:
                        existing_product.shipping_template = new_ship_temp
                        updated = True
                # precautionary_note
                note_raw = base_row.get("precautionary note")
                if note_raw is not None:
                    new_note = clean_str(note_raw)
                    if existing_product.precautionary_note != new_note:
                        existing_product.precautionary_note = new_note
                        updated = True
                # care_instructions
                care_raw = base_row.get("care instructions (if applicable)")
                if care_raw is not None:
                    new_care = clean_str(care_raw)
                    if existing_product.care_instructions != new_care:
                        existing_product.care_instructions = new_care
                        updated = True
                # warranty
                warranty_raw = base_row.get("warranty / guarantee (if applicable)")
                if warranty_raw is not None:
                    new_warranty = clean_str(warranty_raw)
                    if existing_product.warranty != new_warranty:
                        existing_product.warranty = new_warranty
                        updated = True
                # tags
                tag_raw = base_row.get("product tag")
                if tag_raw is not None:
                    new_tags = (
                        [clean_str_with_strip(tag_raw)]
                        if tag_raw and tag_raw.strip()
                        else None
                    )
                    if existing_product.tags != new_tags:
                        existing_product.tags = new_tags
                        updated = True
                # ships_from_location
                ship_loc_raw = base_row.get("ships from location")
                if ship_loc_raw is not None:
                    new_ship_loc = clean_str_with_strip(ship_loc_raw)
                    if existing_product.ships_from_location != new_ship_loc:
                        existing_product.ships_from_location = new_ship_loc
                        updated = True
                # handling_time_days
                hand_raw = base_row.get("handling time (days)")
                if hand_raw is not None:
                    new_hand = to_int(hand_raw)
                    if existing_product.handling_time_days != new_hand:
                        existing_product.handling_time_days = new_hand
                        updated = True
                # fast_dispatch
                fast_raw = base_row.get("fast dispatch")
                if fast_raw is not None:
                    new_fast = to_bool(fast_raw)
                    if existing_product.fast_dispatch != new_fast:
                        existing_product.fast_dispatch = new_fast
                        updated = True
                # free_shipping
                free_raw = base_row.get("free shipping")
                if free_raw is not None:
                    new_free = to_bool(free_raw)
                    if existing_product.free_shipping != new_free:
                        existing_product.free_shipping = new_free
                        updated = True
                # product_condition
                cond_raw = base_row.get("condition")
                if cond_raw is not None:
                    new_cond = clean_str_with_strip(cond_raw)
                    if existing_product.product_condition != new_cond:
                        existing_product.product_condition = new_cond
                        updated = True
                # product_code and ean
                if product_code_raw is not None:
                    if existing_product.product_code != product_code:
                        existing_product.product_code = product_code
                        existing_product.ean = product_code
                        updated = True
                if updated:
                    updated_products += 1
                    await log_activity(
                        db,
                        "product",
                        existing_product.id,
                        "update",
                        {
                            "sku": existing_product.sku,
                            "title": existing_product.title,
                            "event": "import_products_ex",
                        },
                        performed_by=user.get("admin_id"),
                    )
                current_product = existing_product
            else:
                # Check for limited columns before creation
                missing_ess = [
                    f for f in essential_product_fields if base_row.get(f) is None
                ]
                if missing_ess:
                    errors.append(
                        f"Limited columns for product '{product_sku}': missing {missing_ess}. Skipping creation."
                    )
                    continue
                # Create new product
                new_product = Product(
                    sku=product_sku,
                    variation_group_code=variation_group_code,
                    product_code=product_code,
                    ean=product_code,
                    title=clean_str_with_strip(base_row.get("title")),
                    slug=(
                        make_slug(clean_str(base_row.get("title")))
                        if base_row.get("title")
                        else None
                    ),
                    description=clean_str_with_strip(base_row.get("long description")),
                    price=to_decimal(base_row.get("selling price")),
                    product_type=clean_str_with_strip(base_row.get("product_type")),
                    cost_price=to_decimal(base_row.get("cost price")),
                    rrp_price=to_decimal(base_row.get("rrp")),
                    status=clean_str(base_row.get("product status")),
                    weight=to_decimal(base_row.get("package weight (kg)")),
                    length=to_decimal(base_row.get("package length(cms)")),
                    width=to_decimal(base_row.get("package width(cms)")),
                    height=to_decimal(base_row.get("package height(cms)")),
                    bundle_group_code=clean_str_with_strip(
                        base_row.get("bundle / combo indicator")
                    ),
                    unit="kg",
                    supplier=clean_str_with_strip(base_row.get("supplier name")),
                    country_of_origin=clean_str_with_strip(
                        base_row.get("country of origin")
                    ),
                    product_id_type=clean_str_with_strip(
                        base_row.get("product id type")
                    ),
                    asin=None,
                    mpn=None,
                    hs_code=clean_str_with_strip(
                        base_row.get("harmonized code (hs code)")
                    ),
                    category_id=safe_get(
                        category_map, clean_str(base_row.get("category code"))
                    ),
                    brand_id=safe_get(brand_map, clean_str(base_row.get("brand name"))),
                    stock=to_int(base_row.get("stock quantity")),
                    key_features=None,
                    is_battery_required=to_bool(base_row.get("is battery required?")),
                    shipping_template=clean_str_with_strip(
                        base_row.get("shipping template")
                    ),
                    precautionary_note=clean_str(base_row.get("precautionary note")),
                    care_instructions=clean_str(
                        base_row.get("care instructions (if applicable)")
                    ),
                    warranty=clean_str(
                        base_row.get("warranty / guarantee (if applicable)")
                    ),
                    tags=(
                        [clean_str_with_strip(base_row.get("product tag"))]
                        if base_row.get("product tag")
                        else None
                    ),
                    ships_from_location=clean_str_with_strip(
                        base_row.get("ships from location")
                    ),
                    handling_time_days=to_int(base_row.get("handling time (days)")),
                    fast_dispatch=to_bool(base_row.get("fast dispatch")),
                    free_shipping=to_bool(base_row.get("free shipping")),
                    product_condition=clean_str_with_strip(base_row.get("condition")),
                    estimated_shipping_cost=None,
                    product_margin_percent=None,
                    product_margin_amount=None,
                    profit=None,
                )
                db.add(new_product)
                await db.flush()
                await db.refresh(new_product)
                created_products += 1
                await log_activity(
                    db,
                    "product",
                    new_product.id,
                    "create",
                    {
                        "sku": new_product.sku,
                        "title": new_product.title,
                        "event": "import_products_ex",
                    },
                    performed_by=user.get("admin_id"),
                )
                current_product = new_product

            # ---------------- SEO HANDLING ----------------
            seo_keywords = clean_str_with_strip(base_row.get("seo keywords"))
            page_title = clean_str_with_strip(base_row.get("page title"))
            meta_description = clean_str_with_strip(base_row.get("meta description"))
            url_handle = clean_str_with_strip(base_row.get("url handles"))
            canonical_url = clean_str_with_strip(base_row.get("canonical url"))

            if any(
                [seo_keywords, page_title, meta_description, url_handle, canonical_url]
            ):
                seo_rec = await db.execute(
                    select(ProductSEO).where(
                        ProductSEO.product_id == current_product.id
                    )
                )
                existing_seo = seo_rec.scalar_one_or_none()
                if existing_seo:
                    if seo_keywords is not None:
                        existing_seo.meta_keywords = seo_keywords
                    if page_title is not None:
                        existing_seo.page_title = page_title
                    if meta_description is not None:
                        existing_seo.meta_description = meta_description
                    if url_handle is not None:
                        existing_seo.url_handle = url_handle
                    if canonical_url is not None:
                        existing_seo.canonical_url = canonical_url
                else:
                    db.add(
                        ProductSEO(
                            product_id=current_product.id,
                            meta_keywords=seo_keywords,
                            page_title=page_title,
                            meta_description=meta_description,
                            url_handle=url_handle,
                            canonical_url=canonical_url,
                        )
                    )
            # Handle product-level images (replace old)
            await db.execute(
                delete(ProductImage).where(
                    ProductImage.product_id == current_product.id
                )
            )
            product_video_url = clean_str_with_strip(base_row.get("product video link"))
            order = 1
            added_any = False
            for col in image_columns:
                url_raw = base_row.get(col)
                if url_raw is not None:
                    url = clean_str_with_strip(url_raw)
                    is_main = col == "main image url"
                    video = product_video_url if is_main else None
                    if not url and not video:
                        continue
                    img_url = (
                        await download_and_upload_image(
                            url, identifier=current_product.sku
                        )
                        if url
                        else None
                    )
                    db.add(
                        ProductImage(
                            product_id=current_product.id,
                            image_url=img_url,
                            is_main=is_main,
                            video_url=video,
                            image_order=order,
                        )
                    )
                    order += 1
                    added_any = True
            # Fallback video-only
            if not added_any and product_video_url:
                db.add(
                    ProductImage(
                        product_id=current_product.id,
                        image_url=None,
                        is_main=True,
                        video_url=product_video_url,
                        image_order=1,
                    )
                )
            # Handle variants
            for row in sku_rows:
                variant_sku_raw = row.get("sku")
                if variant_sku_raw is None:
                    continue
                variant_sku = clean_str_with_strip(variant_sku_raw)
                if not variant_sku:
                    continue
                # Check ean uniqueness if provided
                ean_raw = row.get("product code")
                ean = clean_str_with_strip(ean_raw) if ean_raw is not None else None
                if ean:
                    result_ean = await db.execute(
                        select(ProductVariant).where(ProductVariant.ean == ean)
                    )
                    existing_ean_check = result_ean.scalar_one_or_none()
                    if existing_ean_check:
                        errors.append(
                            f"Variant EAN '{ean}' already exists. Skipping variant."
                        )
                        continue
                result_var_sku = await db.execute(
                    select(ProductVariant).where(ProductVariant.sku == variant_sku)
                )
                existing_variant = result_var_sku.scalar_one_or_none()
                current_variant = None
                updated_var = False
                if existing_variant:
                    # Update existing variant
                    # title
                    vtitle_raw = row.get("title")
                    if vtitle_raw is not None:
                        new_vtitle = clean_str_with_strip(vtitle_raw)
                        if existing_variant.title != new_vtitle:
                            existing_variant.title = new_vtitle
                            updated_var = True
                    # price
                    vprice_raw = row.get("selling price")
                    if vprice_raw is not None:
                        new_vprice = to_decimal(vprice_raw)
                        if existing_variant.price != new_vprice:
                            existing_variant.price = new_vprice
                            updated_var = True
                    # stock
                    vstock_raw = row.get("stock quantity")
                    vstock_base_raw = base_row.get("stock quantity")
                    if vstock_raw is not None:
                        new_vstock = to_int(vstock_raw)
                        if existing_variant.stock != new_vstock:
                            existing_variant.stock = new_vstock
                            updated_var = True
                    elif vstock_base_raw is not None:
                        new_vstock = to_int(vstock_base_raw)
                        if existing_variant.stock != new_vstock:
                            existing_variant.stock = new_vstock
                            updated_var = True
                    # rrp_price
                    vrrp_raw = row.get("rrp")
                    vrrp_base_raw = base_row.get("rrp")
                    if vrrp_raw is not None:
                        new_vrrp = to_decimal(vrrp_raw)
                        if existing_variant.rrp_price != new_vrrp:
                            existing_variant.rrp_price = new_vrrp
                            updated_var = True
                    elif vrrp_base_raw is not None:
                        new_vrrp = to_decimal(vrrp_base_raw)
                        if existing_variant.rrp_price != new_vrrp:
                            existing_variant.rrp_price = new_vrrp
                            updated_var = True
                    # ean
                    if ean_raw is not None:
                        if existing_variant.ean != ean:
                            existing_variant.ean = ean
                            updated_var = True
                    # ships_from_location
                    vship_loc_raw = row.get("ships from location")
                    vship_loc_base_raw = base_row.get("ships from location")
                    if vship_loc_raw is not None:
                        new_vship = clean_str_with_strip(vship_loc_raw)
                        if existing_variant.ships_from_location != new_vship:
                            existing_variant.ships_from_location = new_vship
                            updated_var = True
                    elif vship_loc_base_raw is not None:
                        new_vship = clean_str_with_strip(vship_loc_base_raw)
                        if existing_variant.ships_from_location != new_vship:
                            existing_variant.ships_from_location = new_vship
                            updated_var = True
                    # handling_time_days
                    vhand_raw = row.get("handling time (days)")
                    vhand_base_raw = base_row.get("handling time (days)")
                    if vhand_raw is not None:
                        new_vhand = to_int(vhand_raw)
                        if existing_variant.handling_time_days != new_vhand:
                            existing_variant.handling_time_days = new_vhand
                            updated_var = True
                    elif vhand_base_raw is not None:
                        new_vhand = to_int(vhand_base_raw)
                        if existing_variant.handling_time_days != new_vhand:
                            existing_variant.handling_time_days = new_vhand
                            updated_var = True
                    # dimensions from base
                    vlength_raw = base_row.get("package length(cms)")
                    if vlength_raw is not None:
                        new_vlength = to_decimal(vlength_raw)
                        if existing_variant.length != new_vlength:
                            existing_variant.length = new_vlength
                            updated_var = True
                    vwidth_raw = base_row.get("package width(cms)")
                    if vwidth_raw is not None:
                        new_vwidth = to_decimal(vwidth_raw)
                        if existing_variant.width != new_vwidth:
                            existing_variant.width = new_vwidth
                            updated_var = True
                    vheight_raw = base_row.get("package height(cms)")
                    if vheight_raw is not None:
                        new_vheight = to_decimal(vheight_raw)
                        if existing_variant.height != new_vheight:
                            existing_variant.height = new_vheight
                            updated_var = True
                    vweight_raw = base_row.get("package weight (kg)")
                    if vweight_raw is not None:
                        new_vweight = to_decimal(vweight_raw)
                        if existing_variant.weight != new_vweight:
                            existing_variant.weight = new_vweight
                            updated_var = True
                    # bundle_group_code from base
                    vbundle_raw = base_row.get("bundle / combo indicator")
                    if vbundle_raw is not None:
                        new_vbundle = clean_str_with_strip(vbundle_raw)
                        if existing_variant.bundle_group_code != new_vbundle:
                            existing_variant.bundle_group_code = new_vbundle
                            updated_var = True
                    # free_shipping from base
                    vfree_raw = base_row.get("free shipping")
                    if vfree_raw is not None:
                        new_vfree = to_bool(vfree_raw)
                        if existing_variant.free_shipping != new_vfree:
                            existing_variant.free_shipping = new_vfree
                            updated_var = True
                    # fast_dispatch from base
                    vfast_raw = base_row.get("fast dispatch")
                    if vfast_raw is not None:
                        new_vfast = to_bool(vfast_raw)
                        if existing_variant.fast_dispatch != new_vfast:
                            existing_variant.fast_dispatch = new_vfast
                            updated_var = True
                    # hs_code from base
                    vhs_raw = base_row.get("harmonized code (hs code)")
                    if vhs_raw is not None:
                        new_vhs = clean_str_with_strip(vhs_raw)
                        if existing_variant.hs_code != new_vhs:
                            existing_variant.hs_code = new_vhs
                            updated_var = True
                    current_variant = existing_variant
                    if updated_var:
                        updated_variants += 1
                else:
                    # Check for limited columns before variant creation
                    missing_ess_var = [
                        f for f in essential_variant_fields if row.get(f) is None
                    ]
                    if missing_ess_var:
                        errors.append(
                            f"Limited columns for variant '{variant_sku}': missing {missing_ess_var}. Skipping creation."
                        )
                        continue
                    # Create new variant
                    new_variant = ProductVariant(
                        product_id=current_product.id,
                        title=clean_str_with_strip(row.get("title")),
                        sku=variant_sku,
                        ean=ean,
                        price=to_decimal(row.get("selling price")),
                        stock=to_int(
                            row.get("stock quantity") or base_row.get("stock quantity")
                        ),
                        cost_price=0,
                        ships_from_location=clean_str_with_strip(
                            row.get("ships from location")
                            or base_row.get("ships from location")
                        ),
                        handling_time_days=to_int(
                            row.get("handling time (days)")
                            or base_row.get("handling time (days)")
                        ),
                        length=to_decimal(base_row.get("package length(cms)")),
                        width=to_decimal(base_row.get("package width(cms)")),
                        height=to_decimal(base_row.get("package height(cms)")),
                        weight=to_decimal(base_row.get("package weight (kg)")),
                        bundle_group_code=clean_str_with_strip(
                            base_row.get("bundle / combo indicator")
                        ),
                        estimated_shipping_cost=None,
                        product_margin_percent=None,
                        product_margin_amount=None,
                        profit=None,
                        rrp_price=to_decimal(row.get("rrp") or base_row.get("rrp")),
                        free_shipping=to_bool(base_row.get("free shipping")),
                        fast_dispatch=to_bool(base_row.get("fast dispatch")),
                        hs_code=clean_str_with_strip(
                            base_row.get("harmonized code (hs code)")
                        ),
                    )
                    db.add(new_variant)
                    await db.flush()
                    await db.refresh(new_variant)
                    current_variant = new_variant
                    created_variants += 1
                # Handle attributes (replace old)
                await db.execute(
                    delete(Attribute).where(Attribute.variant_id == current_variant.id)
                )
                max_options = 10
                for i in range(1, max_options + 1):
                    option_col = f"variant option {i} (e.g., colour, size, style)"
                    value_col = f"variant values {i}"
                    name_raw = row.get(option_col)
                    value_raw = row.get(value_col)
                    if name_raw is not None and value_raw is not None:
                        name = clean_str_with_strip(name_raw)
                        value = clean_str_with_strip(value_raw)
                        if name and value:
                            db.add(
                                Attribute(
                                    name=name,
                                    value=value,
                                    variant_id=current_variant.id,
                                )
                            )
                # Handle variant-level images (replace old)
                await db.execute(
                    delete(ProductVariantImage).where(
                        ProductVariantImage.variant_id == current_variant.id
                    )
                )
                order = 1
                added_any = False
                for col in image_columns:
                    url_raw = row.get(col)
                    if url_raw is not None:
                        url = clean_str_with_strip(url_raw)
                        is_main = col == "main image url"
                        video = None  # No per-variant video in schema
                        if not url and not video:
                            continue
                        img_url = (
                            await download_and_upload_image(
                                url, identifier=current_variant.sku
                            )
                            if url
                            else None
                        )
                        db.add(
                            ProductVariantImage(
                                variant_id=current_variant.id,
                                image_url=img_url,
                                is_main=is_main,
                                video_url=video,
                                image_order=order,
                            )
                        )
                        order += 1
                        added_any = True
                # No fallback for variants as video is product-level
    return {
        "created_products": created_products,
        "updated_products": updated_products,
        "created_variants": created_variants,
        "updated_variants": updated_variants,
        "errors": errors,
    }


from sqlalchemy.future import select
from sqlalchemy.orm import sessionmaker


@router.post("/import-reviews", status_code=status.HTTP_201_CREATED)
async def import_reviews(
    file: Annotated[UploadFile, File(...)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    contents = await file.read()
    filename = file.filename
    ext = filename.split(".")[-1].lower()

    rows = []

    MANDATORY_COLUMNS = {
        "sku",
        "customer name",
        "star rating",
        "title",
        "customer comment",
        "date",
    }

    if ext == "csv":
        text = contents.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))

        headers = [(h.strip().lower() if h else "") for h in reader.fieldnames]
        missing_cols = [col for col in MANDATORY_COLUMNS if col not in headers]
        if missing_cols:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing mandatory columns: {missing_cols}. Please use the standard format.",
            )

        reader.fieldnames = headers
        for row in reader:
            normalized_row = {
                (k.strip().lower() if k else ""): v for k, v in row.items()
            }
            if all(not v for v in normalized_row.values()):
                continue
            rows.append(normalized_row)

    elif ext in ("xls", "xlsx"):
        wb = load_workbook(BytesIO(contents), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        try:
            # Extract and validate headers
            headers = [
                str(h).strip().lower() if h is not None else "" for h in next(rows_iter)
            ]
            missing_cols = [col for col in MANDATORY_COLUMNS if col not in headers]
            if missing_cols:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Missing mandatory columns: {missing_cols}. Please use the standard format.",
                )
        except StopIteration:
            return {"created_reviews": 0, "errors": ["File is empty"]}

        for r in rows_iter:
            if not r or all(cell is None or str(cell).strip() == "" for cell in r):
                continue

            row_dict = {}
            for i, header in enumerate(headers):
                if header:
                    row_dict[header] = r[i] if i < len(r) else None
            rows.append(row_dict)

    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload .csv, .xls, or .xlsx.",
        )

    created_reviews = 0
    errors = []
    affected_product_ids = set()

    image_columns = ["image 1", "image 2", "image 3", "image 4"]

    for idx, row in enumerate(rows):
        # Map columns: SKU, Customer Name, Star Rating, Customer Comment
        sku = row.get("sku")
        if not sku:
            errors.append(f"Row {idx + 2}: Missing SKU")
            continue

        sku = str(sku).strip()

        # Find product and/or variant
        product_id = None
        variant_id = None

        # 1. Check ProductVariant table FIRST (specific variant match)
        result_var = await db.execute(
            select(ProductVariant).where(ProductVariant.sku == sku)
        )
        variant = result_var.scalar_one_or_none()

        if variant:
            product_id = variant.product_id
            variant_id = variant.id
        else:
            # 2. Check Product table (product match)
            result = await db.execute(select(Product).where(Product.sku == sku))
            product = result.scalar_one_or_none()

            if product:
                product_id = product.id
                # validation: ensure this product doesn't have variants?
                # If product has variants, usually review should attach to a variant.
                # But requirements say "if a product dont have variant and sku is product only save for that"
                # If product HAS variants but user gave product SKU, do we save to product?
                # Yes, per "sku is product only save for that" logic implies fallback to product level.

        if not product_id:
            errors.append(
                f"Row {idx + 2}: Product or Variant with SKU '{sku}' not found"
            )
            continue

        customer_name = row.get("customer name")
        if not customer_name:
            customer_name = "Anonymous"
        else:
            customer_name = str(customer_name).strip()

        star_rating_raw = row.get("star rating")
        rating = None
        if star_rating_raw is not None:
            try:
                rating = float(star_rating_raw)
                if not (1 <= rating <= 5):
                    errors.append(f"Row {idx + 2}: Invalid rating '{rating}'")
                    continue
            except ValueError:
                errors.append(f"Row {idx + 2}: Invalid rating '{star_rating_raw}'")
                continue

        comment = row.get("customer comment")
        if comment:
            comment = str(comment).strip()

        title = row.get("title")
        if title:
            title = str(title).strip()
        else:
            title = "Product Review"

        date_raw = row.get("date")
        created_at = None
        if date_raw:
            try:
                import pandas as pd

                if isinstance(date_raw, datetime):
                    created_at = date_raw
                else:
                    parsed_date = pd.to_datetime(date_raw)
                    if pd.notnull(parsed_date):
                        created_at = parsed_date.to_pydatetime()
            except Exception:
                pass

        tasks = []
        for col in image_columns:
            val = row.get(col)
            if val:
                url = str(val).strip()
                if url and url.lower() not in ["nan", "none", ""]:
                    tasks.append(download_and_upload_image(url, identifier="review"))

        images = await asyncio.gather(*tasks)
        images = [img for img in images if img][:4]
        print("images---------------------", images)

        review_kwargs = {
            "product_id": product_id,
            "variant_id": variant_id,
            "product_identifier": sku,
            "reviewer_name": customer_name,
            "rating": rating,
            "title": title,
            "comment": comment,
            "review_type": "customer",
            "is_verified_purchase": False,
            "is_official": False,
            "images": images if images else None,
        }

        if created_at:
            review_kwargs["created_at"] = created_at

        # Create Review
        new_review = Review(**review_kwargs)

        db.add(new_review)
        affected_product_ids.add(product_id)
        created_reviews += 1

    await db.commit()

    for pid in affected_product_ids:
        await update_review_stats(db, pid)

    return {
        "created_reviews": created_reviews,
        "errors": errors,
        "message": "Import completed",
    }


@router.get("/export-reviews", response_class=StreamingResponse)
async def export_reviews(db: Annotated[AsyncSession, Depends(get_db)]):
    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews"

    headers = [
        "SKU",
        "Customer Name",
        "Star Rating",
        "Title",
        "Customer Comment",
        "Date",
        "Image 1",
        "Image 2",
        "Image 3",
        "Image 4",
    ]
    ws.append(headers)

    # Style headers (bold)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Fetch reviews with product relation
    result = await db.execute(
        select(Review).options(
            selectinload(Review.product), selectinload(Review.variant)
        )
    )
    reviews = result.scalars().all()

    for review in reviews:
        # Determine SKU: prioritize live Variant SKU, then live Product SKU, then fallback to snapshot
        sku = review.product_identifier
        if review.variant and review.variant.sku:
            sku = review.variant.sku
        elif review.product and review.product.sku:
            sku = review.product.sku

        images = review.images or []

        ws.append(
            [
                sku,
                review.reviewer_name,
                review.rating,
                review.title,
                review.comment,
                (
                    review.created_at.replace(tzinfo=None)
                    if hasattr(review, "created_at") and review.created_at
                    else None
                ),
                images[0] if len(images) > 0 else None,
                images[1] if len(images) > 1 else None,
                images[2] if len(images) > 2 else None,
                images[3] if len(images) > 3 else None,
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reviews_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/download-reviews-template", response_class=StreamingResponse)
async def download_reviews_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews Template"

    # Headers
    headers = [
        "SKU",
        "Customer Name",
        "Star Rating",
        "Title",
        "Customer Comment",
        "Date",
        "Image 1",
        "Image 2",
        "Image 3",
        "Image 4",
    ]
    ws.append(headers)

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Example row (IMPORTANT)
    ws.append(
        [
            "ABC123",
            "John Doe",
            5,
            "Great product",
            "Loved it",
            datetime.now().strftime("%Y-%m-%d"),
            "https://example.com/img1.jpg",
            "https://example.com/img2.jpg",
            "https://example.com/img3.jpg",
            "https://example.com/img4.jpg",
        ]
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reviews_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
