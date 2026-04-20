# apis/v1/promotion.py
from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from models.promotions import Promotion
from schemas.promotions import PromotionCreate, PromotionResponse
from deps import get_db
from utils.promotions_utils import validate_reference_exists
from datetime import datetime
from fastapi.responses import FileResponse
from openpyxl import Workbook
from models.product import Product
from openpyxl.styles import Font
from sqlalchemy.future import select
from schemas.promotions import ProductIdRequest, PaginatedPromotions
from openpyxl import load_workbook
from typing import Annotated
from fastapi import Query
from fastapi import Query
from sqlalchemy import func
from utils.activity_logger import log_activity

router = APIRouter()


@router.get(
    "/get-promotion/{product_id}",
    responses={
        404: {"description": "No active promotions found"},
    },
)
async def get_promotions_by_product(
    product_id: str, db: Annotated[AsyncSession, Depends(get_db)]
):
    result = await db.execute(
        select(Promotion).where(
            Promotion.reference_id == product_id, Promotion.status == "active"
        )
    )
    promotions = result.scalars().all()

    if not promotions:
        raise HTTPException(status_code=404, detail="No active promotions found")

    return [
        {
            "id": promo.id,
            "offer_name": promo.offer_name,
            "discount_type": promo.discount_type,
            "discount_value": promo.discount_value,
            "discount_percentage": promo.discount_percentage,
            "max_discount_amount": promo.max_discount_amount,
            "description": promo.description,
            "start_date": promo.start_date,
            "end_date": promo.end_date,
        }
        for promo in promotions
    ]


@router.post(
    "/create-promotions",
    response_model=PromotionResponse,
    responses={
        400: {"description": "An active promotion already exists for this entity."},
    },
)
async def create_promotion(
    data: PromotionCreate, db: Annotated[AsyncSession, Depends(get_db)]
):
    """Create a new promotional offer for product or category with reference validation."""

    await validate_reference_exists(db, data.offer_type, data.reference_id)

    existing = await db.execute(
        select(Promotion)
        .where(Promotion.offer_type == data.offer_type)
        .where(Promotion.reference_id == data.reference_id)
        .where(Promotion.status == "active")
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"An active promotion already exists for this {data.offer_type}.",
        )

    new_promo = Promotion(
        offer_name=data.name,
        description=data.description,
        discount_type=data.discount_type,
        discount_value=data.discount_value,
        discount_percentage=data.discount_percentage,
        max_discount_amount=data.max_discount_amount,
        offer_type=data.offer_type,
        reference_id=data.reference_id,
        start_date=data.start_date,
        end_date=data.end_date,
    )

    db.add(new_promo)
    await db.flush()
    await log_activity(
        db,
        entity_type="promotion",
        entity_id=new_promo.id,
        action="create",
        details={"name": new_promo.offer_name},
        performed_by="admin"
    )
    await db.commit()
    await db.refresh(new_promo)

    return new_promo


@router.get("/list-promotions")
async def list_promotions(
    db: Annotated[AsyncSession, Depends(get_db)],
    page: Annotated[int, Query(ge=1)] = 1,
    size: Annotated[int, Query(ge=1, le=100)] = 10,
):
    # --- total count ---
    total_result = await db.execute(select(func.count(Promotion.id)))
    total = total_result.scalar() or 0

    # --- fetch page slice ---
    result = await db.execute(
        select(Promotion)
        .order_by(Promotion.created_at.desc())
        .offset((page - 1) * size)
        .limit(size)
    )

    promotions = result.scalars().all()

    # dynamically update status
    for promo in promotions:
        promo.status = promo.current_status

    return {
        "total": total,
        "page": page,
        "limit": size,
        "pages": (total + size - 1) // size,
        "data": promotions,
    }


@router.delete(
    "/delete-promotions/{promotion_id}",
    responses={
        404: {"description": "Promotion not found"},
    },
)
async def delete_promotion(
    promotion_id: str, db: Annotated[AsyncSession, Depends(get_db)]
):
    """Delete a promotion by ID."""
    result = await db.execute(select(Promotion).where(Promotion.id == promotion_id))
    promo = result.scalar_one_or_none()

    if not promo:
        raise HTTPException(status_code=404, detail="Promotion not found")

    await db.delete(promo)
    await db.commit()
    return {"message": "Promotion deleted successfully"}


promotions_headers = [
    "Product Identifier / SKU",
    "Event Name",
    "Promotional Discount %",
    "Selling Price",
    "Discounted Price",
    "Sale Start Date",
    "Sale End Date",
]


@router.post("/promotions-export-template")
async def export_promotion_template(
    payload: ProductIdRequest, db: Annotated[AsyncSession, Depends(get_db)]
):

    result = await db.execute(
        select(Product.id, Product.sku).where(Product.id.in_(payload.product_ids))
    )

    products = result.all()  # <-- list of tuples (id, sku)

    wb = Workbook()
    ws = wb.active
    ws.title = "Promotion Template"

    ws.append(promotions_headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for _, sku in products:
        ws.append(
            [
                sku,
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"/tmp/promo_template_{timestamp}.xlsx"
    wb.save(filename)

    return FileResponse(
        filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"promotion_template_{timestamp}.xlsx",
    )


@router.post(
    "/promotions-import-template",
    responses={
        400: {"description": "Invalid file format or template."},
    },
)
async def import_promotions(
    file: Annotated[UploadFile, File(...)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    # --- Validate file type ---
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Invalid file format")

    wb = load_workbook(file.file, data_only=True)
    ws = wb.active

    # --- Validate headers ---
    sheet_headers = [c.value for c in ws[1]]
    if sheet_headers != promotions_headers:
        raise HTTPException(
            status_code=400,
            detail="Invalid template. Please use the exported promotion template.",
        )

    created = 0
    skipped = 0
    errors = []

    # --- Process rows ---
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        (
            sku,
            event_name,
            discount_percent,
            selling_price,
            discounted_price,
            start_date,
            end_date,
        ) = row

        # Skip empty rows
        if not sku:
            skipped += 1
            continue

        # Validate required fields
        if not event_name or not start_date or not end_date:
            errors.append(f"Row {idx}: Missing required promotion fields")
            continue

        # Parse dates (Excel can store dates as datetime already)
        if isinstance(start_date, str):
            start_date = datetime.fromisoformat(start_date)

        if isinstance(end_date, str):
            end_date = datetime.fromisoformat(end_date)

        if start_date >= end_date:
            errors.append(f"Row {idx}: Start date must be before end date")
            continue

        # --- Find product by SKU ---
        result = await db.execute(select(Product).where(Product.sku == sku))
        product = result.scalar_one_or_none()

        if not product:
            errors.append(f"Row {idx}: Product not found for SKU {sku}")
            continue

        # --- Create promotion ---
        promo = Promotion(
            offer_name=event_name,
            offer_type="product",
            reference_id=product.id,
            discount_type="percentage" if discount_percent else "fixed",
            discount_percentage=discount_percent,
            discount_value=discounted_price if discounted_price else 0,
            original_price=selling_price,
            discounted_price=discounted_price,
            start_date=start_date,
            end_date=end_date,
            status="active",
        )

        db.add(promo)
        created += 1

    await db.commit()

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }
