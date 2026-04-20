import io
import pandas as pd
import re
import unicodedata
import csv
import logging
import asyncio
import os
import redis
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any, Tuple, Optional, Literal, Set
from io import StringIO, BytesIO
import secrets
import random
import string
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, delete, update, or_, func, and_, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from dotenv import load_dotenv

from models.product import Product, ProductVariant, ProductImage, ProductVariantImage, Attribute
from models.warehouse import ProductStock
from models.category import Category
from models.brand import Brand
from models.seo import ProductSEO
from models.b_tasks import BackgroundTask
from deps import get_db
from schemas.import_export import (
    ProductImportRow,
    format_pydantic_error,
    BulkTemplateRequest,
)
from service.redis import get_redis_url
from urllib.parse import urlparse
from utils.image_handler import download_and_upload_image
from utils.activity_logger import log_activity
from sqlalchemy.exc import IntegrityError, SQLAlchemyError, PendingRollbackError
from database import create_sessionmaker


load_dotenv()

logger = logging.getLogger(__name__)


def safe_get(mapping: dict | None, key: str):
    """Safely get a value from a dict that may be None.
    Returns None if mapping is None or key not present."""
    if mapping is None:
        return None
    return mapping.get(key)



def get_first(row: dict, keys: list[str]) -> str | None:
    """Return the first non-None value for given possible keys in a dict."""
    for k in keys:
        if k in row:
            val = row.get(k)
            if val is not None:
                return val
    return None

def create_redis_client():
    host = os.getenv("REDIS_HOST", "10.192.0.2")
    port = os.getenv("REDIS_PORT", 6379)
    print(host, port, "this is redis host and port")
    return redis.Redis(host=host, port=port, db=0)


redis_client = create_redis_client()


def make_slug(value: str | None) -> str:
    if not value:
        return ""
    value = unicodedata.normalize("NFKD", value)
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = value.strip("-")
    value = re.sub(r"-+", "-", value)
    return value


async def generate_unique_product_code(db: AsyncSession, existing_codes: Set[str] = None) -> str:
    """Generate a unique 10-digit alphanumeric product code."""
    chars = string.ascii_uppercase + string.digits
    while True:
        code = "".join(secrets.choice(chars) for _ in range(10))
        if existing_codes is not None:
            if code not in existing_codes:
                existing_codes.add(code)
                return code
        else:
            # Check if code already exists in DB
            result = await db.execute(select(Product).where(Product.unique_code == code))
            if not result.scalar_one_or_none():
                return code


async def validate_brand_and_category(db, rows, Brand, Category):
    brand_names: Set[str] = set()
    category_codes: Set[str] = set()
    for r in rows:
        b = r.get("brand name")
        c = r.get("category code")
        if b:
            brand_names.add(str(b).strip().lower())
        if c:
            category_codes.add(str(c).strip().lower())
    errors = []
    brand_map = {}
    category_map = {}
    await db.commit()
    # -------- BRANDS --------
    if brand_names:
        q = await db.execute(
            select(Brand.id, Brand.name).where(func.lower(Brand.name).in_(brand_names))
        )
        rows_found = q.all()
        brand_map = {name.lower(): bid for bid, name in rows_found}
        missing = brand_names - set(brand_map.keys())
        for m in missing:
            errors.append(f"Brand name '{m}' does not exist in database.")
    # -------- CATEGORIES --------
    if category_codes:
        q = await db.execute(
            select(Category.id, Category.category_code).where(
                func.lower(Category.category_code).in_(category_codes)
            )
        )
        # existing = {row[0] for row in q.all()}
        rows_found = q.all()
        category_map = {c_code.lower(): cid for cid, c_code in rows_found}
        missing = category_codes - set(category_map.keys())
        for m in missing:
            errors.append(f"Category code '{m}' does not exist in database.")
    if errors:
        return False, errors, brand_map, category_map
    return True, [], brand_map, category_map


def clean_str_with_strip(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
    else:
        s = str(value).strip()
    return s if s else None


def clean_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
    else:
        s = str(value).strip()
    return s.lower() if s else None


def to_int(value: Any) -> int | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(float(s))  # handles "10.0" too
    except ValueError:
        return None


def to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def to_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if not s:
        return None
    if s in ("yes", "y", "true", "1"):
        return True
    if s in ("no", "n", "false", "0"):
        return False
    return None


# apis/v1/product_import_export.py
import io
import pandas as pd
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
import re
import unicodedata
from models.product import *
from deps import get_db
from fastapi.responses import StreamingResponse
import csv
from datetime import datetime
from fastapi import status
from sqlalchemy.orm import selectinload
from models.product import ProductVariant
from models.category import Category
from models.brand import Brand
from typing import List, Dict, Any
from io import StringIO, BytesIO
from openpyxl import load_workbook
from decimal import Decimal
from typing import Literal
from fastapi import APIRouter, Query, Depends
from sqlalchemy import select, delete
from typing import Tuple
from openpyxl import Workbook
from typing import Optional, List, Literal
from openpyxl.styles import Font
from schemas.import_export import (
    ProductImportRow,
    format_pydantic_error,
    BulkTemplateRequest,
)
from .utils import *

# from utils.import_export import _read_file_rows
from models.product import ProductImage, ProductVariantImage
from urllib.parse import urlparse
from sqlalchemy import func
from typing import Dict, List, Any
from fastapi import UploadFile, File, Depends, HTTPException, status
from sqlalchemy import select, delete
from sqlalchemy.exc import IntegrityError

router = APIRouter()


import logging
import asyncio

logger = logging.getLogger(__name__)


from fastapi import HTTPException, status, UploadFile
from typing import List, Dict, Any
import csv
from openpyxl import load_workbook
from io import BytesIO, StringIO

from typing import List, Dict, Any
from io import BytesIO, StringIO
import csv
from openpyxl import load_workbook
from fastapi import HTTPException, status
import os
from database import create_sessionmaker
def _read_file_rows(file: str) -> List[Dict[str, Any]]:
    
    ext = file.split(".")[-1].lower()

    with open(file, "rb") as f:
        contents = f.read()

    if ext == "csv":
        text = contents.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        reader.fieldnames = [
            (h.strip().lower() if h is not None else "") for h in reader.fieldnames
        ]
        rows: List[Dict[str, Any]] = []
        for row in reader:
            normalized_row = {(k.strip().lower() if k else ""): v for k, v in row.items()}
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in normalized_row.values()):
                continue
            rows.append(normalized_row)
        return rows

    if ext in ("xls", "xlsx"):
        wb = load_workbook(BytesIO(contents), read_only=True)
        sheet_name = None
        for name in wb.sheetnames:
            if name.strip().lower() == "products template":
                sheet_name = name
                break
        ws = wb[sheet_name] if sheet_name else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip().lower() if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []

        rows: List[Dict[str, Any]] = []
        for r in rows_iter:
            if r is None or all((cell is None or str(cell).strip() == "") for cell in r):
                continue
            row_dict: Dict[str, Any] = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                value = r[i] if i < len(r) else None
                row_dict[header] = value
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_dict.values()):
                continue
            rows.append(row_dict)
        return rows

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Unsupported file type. Please upload .csv, .xls, or .xlsx.",
    )


MANDATORY_FIELDS = {
    "title": "title",
    "sku":"sku",
    "selling price": "price",
    "product status": "status",
    "category code": "category",
}

IMAGE_COLUMNS = [
    "main image url", "lifestyle image url",
    "image 1 url", "image 2 url", "image 3 url", "image 4 url",
    "image 5 url", "image 6 url", "image 7 url", "image 8 url",
    "image 9 url", "image 10 url",
]

# TRYING UPDATE

def validate_mandatory_fields(row, row_no):
    missing = []
    for col, label in MANDATORY_FIELDS.items():
        val = row.get(col)
        if val is None or str(val).strip() == "":
            missing.append(label)

    if missing:
        return f"Row {row_no}: Missing mandatory fields: {', '.join(missing)}"
    return None



async def process_batch(db, rows_with_idx, brand_map, category_map, existing_codes, downloaded_images, admin_id="system"):
    created_products = 0
    updated_products = 0
    created_variants = 0
    updated_variants = 0
    deleted_products = 0
    deleted_variants = 0
    errors = []

    grouped = {}
    
    # -----------------------------
    # GROUP ROWS BY PRODUCT KEY
    # -----------------------------
    for row_no, row in rows_with_idx:
        sku_raw = row.get("sku")
        parent_code_raw = row.get("parent sku / variation group code")
        if parent_code_raw:
            group_key = clean_str_with_strip(parent_code_raw)
        else:
            if not sku_raw:
                errors.append(
                    f"Row {row_no}: No SKU or parent code provided. Skipping."
                )
                continue
            group_key = clean_str_with_strip(sku_raw)
        grouped.setdefault(group_key, []).append((row_no, row))

    # -----------------------------
    # PRE-FETCH DATA FOR THE BATCH
    # -----------------------------
    all_skus = set()
    all_variant_skus = set()
    for row_no, row in rows_with_idx:
        sku = clean_str_with_strip(row.get("sku"))
        parent = clean_str_with_strip(row.get("parent sku / variation group code"))
        if sku: all_skus.add(sku)
        if parent: all_skus.add(parent)
        # Also variant SKUs
        if sku and parent: all_variant_skus.add(sku)

    async with db.begin():
        # 🔹 3. Set aggressive session timeouts per batch (in case of fresh connection)
        if db.bind.dialect.name == "mysql":
            await db.execute(text("SET SESSION wait_timeout = 600"))
            await db.execute(text("SET SESSION interactive_timeout = 600"))
        # Fetch products
        products_map = {}
        if all_skus:
            p_res = await db.execute(select(Product).where(Product.sku.in_(all_skus)))
            for p in p_res.scalars().all():
                products_map[p.sku] = p

        # Fetch variants
        variants_map = {}
        if all_variant_skus:
            v_res = await db.execute(select(ProductVariant).where(ProductVariant.sku.in_(all_variant_skus)))
            for v in v_res.scalars().all():
                variants_map[v.sku] = v

        # Fetch SEO
        seo_map = {}
        if products_map:
            p_ids = [p.id for p in products_map.values()]
            s_res = await db.execute(select(ProductSEO).where(ProductSEO.product_id.in_(p_ids)))
            for s in s_res.scalars().all():
                seo_map[s.product_id] = s

        # -----------------------------
        # PROCESS EACH PRODUCT GROUP
        # -----------------------------
        for product_key, grouped_rows in grouped.items():
            row_no, base_row = grouped_rows[0]
            try:
                if clean_str_with_strip(base_row.get("action")) is None:
                    action = 'update'
                else:
                    action = clean_str_with_strip(base_row.get("action")).lower()

                # -----------------------------
                # FETCH PRODUCT
                # -----------------------------
                product_sku = clean_str_with_strip(base_row.get("sku"))
                product_parent_code = clean_str_with_strip(base_row.get("parent sku / variation group code"))
                
                if product_parent_code:
                    product = products_map.get(product_parent_code)
                else:
                    product = products_map.get(product_sku)

                # -----------------------------
                # DELETE PRODUCT
                # -----------------------------
                if action.lower() == "delete":
                    if product:
                        await db.execute(delete(ProductSEO).where(ProductSEO.product_id == product.id))
                        await db.execute(delete(ProductStock).where(ProductStock.product_id == product.id))
                        await db.execute(delete(ProductVariant).where(ProductVariant.product_id == product.id))
                        await db.execute(delete(ProductImage).where(ProductImage.product_id == product.id))
                        await db.execute(delete(Product).where(Product.id == product.id))
                        deleted_products += 1
                        await log_activity(db, "product", product.id, "delete", {"sku": product.sku, "title": product.title, "event": "bulk_import"}, performed_by=admin_id)
                    continue

                # -----------------------------
                # UPDATE PRODUCT
                # -----------------------------
                if product:
                    updated = False

                    def update_if_changed(field, value):
                        nonlocal updated
                        if value is not None and getattr(product, field) != value:
                            setattr(product, field, value)
                            updated = True

                    title = clean_str_with_strip(base_row.get("title"))
                    if title:
                        update_if_changed("title", title)
                        product.slug = make_slug(title)

                    update_if_changed("description", clean_str_with_strip(base_row.get("long description")))
                    update_if_changed("ships_from_location", clean_str_with_strip(base_row.get("ships from location")))
                    update_if_changed("product_condition", clean_str_with_strip(base_row.get("condition")))
                    update_if_changed("rrp_price", to_decimal(base_row.get("rrp")))
                    update_if_changed("price", to_decimal(base_row.get("selling price")))
                    update_if_changed("cost_price", to_decimal(base_row.get("cost price")))
                    update_if_changed("product_type", clean_str_with_strip(base_row.get("product_type")))
                    update_if_changed("status", clean_str(base_row.get("product status")))
                    update_if_changed("weight", to_decimal(base_row.get("package weight (kg)")))
                    update_if_changed("length", to_decimal(base_row.get("package length(cms)")))
                    update_if_changed("width", to_decimal(base_row.get("package width(cms)")))
                    update_if_changed("height", to_decimal(base_row.get("package height(cms)")))
                    update_if_changed("bundle_group_code", clean_str_with_strip(base_row.get("bundle / combo indicator")))
                    update_if_changed("supplier", clean_str_with_strip(base_row.get("supplier name")))
                    update_if_changed("country_of_origin", clean_str_with_strip(base_row.get("country of origin")))
                    update_if_changed("product_id_type", clean_str_with_strip(base_row.get("product id type")))
                    update_if_changed("hs_code", clean_str_with_strip(base_row.get("harmonized code (hs code)")))
                    update_if_changed("stock", to_int(base_row.get("stock quantity")))
                    update_if_changed("handling_time_days", to_int(base_row.get("handling time (days)")))
                    update_if_changed("fast_dispatch", to_bool(base_row.get("fast dispatch")))
                    update_if_changed("free_shipping", to_bool(base_row.get("free shipping")))
                    update_if_changed("precautionary_note", clean_str_with_strip(base_row.get("precautionary note")))
                    update_if_changed("care_instructions", clean_str_with_strip(get_first(base_row, ["care instructions (if applicable)", "care instructions (if any)"])))
                    update_if_changed("warranty", clean_str_with_strip(get_first(base_row, ["warranty / guarantee (if applicable)", "warranty / guarantee"])))
                    
                    p_code = clean_str_with_strip(base_row.get("product code"))
                    if p_code and product.product_code != p_code:
                        product.product_code = p_code
                        updated = True
                    
                    parent_code_raw = base_row.get("parent sku / variation group code")
                    variation_group_code = clean_str_with_strip(parent_code_raw) if parent_code_raw else None
                    if product.variation_group_code != variation_group_code:
                        product.variation_group_code = variation_group_code
                        updated = True
                    
                    has_variants = any(clean_str_with_strip(r.get("parent sku / variation group code")) is not None for _, r in grouped_rows)
                    new_ean = None if has_variants else clean_str_with_strip(base_row.get("product code"))
                    if product.ean != new_ean:
                        product.ean = new_ean
                        updated = True
                    
                    product_tag = clean_str_with_strip(base_row.get("product tag"))
                    if product_tag:
                        new_tags = [product_tag]
                        if product.tags != new_tags:
                            product.tags = new_tags
                            updated = True

                    category_code = clean_str(base_row.get("category code"))
                    if category_code:
                        cid = category_map.get(category_code)
                        if cid: update_if_changed("category_id", cid)
                    
                    brand_name = clean_str(base_row.get("brand name"))
                    if brand_name:
                        bid = brand_map.get(brand_name)
                        if bid: update_if_changed("brand_id", bid)

                    if updated:
                        updated_products += 1
                        await log_activity(db, "product", product.id, "update", {"sku": product.sku, "title": product.title, "event": "bulk_import"}, performed_by=admin_id)

                # -----------------------------
                # CREATE PRODUCT
                # -----------------------------
                else:
                    price = to_decimal(base_row.get("selling price"))
                    if price is None:
                        errors.append(f"Row {row_no}: Invalid price for SKU {base_row.get('sku')}")
                        continue

                    category_id = category_map.get(clean_str(base_row.get("category code")))
                    brand_id = brand_map.get(clean_str(base_row.get("brand name")))
                    
                    if not category_id:
                        errors.append(f"Row {row_no}: Category code '{base_row.get('category code')}' not found. Skipping product.")
                        continue
                    if not brand_id:
                        errors.append(f"Row {row_no}: Brand name '{base_row.get('brand name')}' not found. Skipping product.")
                        continue

                    parent_code_raw = base_row.get("parent sku / variation group code")
                    variation_group_code = clean_str_with_strip(parent_code_raw) if parent_code_raw else None
                    unique_code = await generate_unique_product_code(db, existing_codes=existing_codes)

                    product = Product(
                        unique_code=unique_code,
                        sku=variation_group_code if variation_group_code else clean_str_with_strip(base_row.get("sku")),
                        variation_group_code=variation_group_code,
                        title=clean_str_with_strip(base_row.get("title")),
                        slug=make_slug(clean_str(base_row.get("title"))),
                        description=clean_str_with_strip(base_row.get("long description")),
                        price=price,
                        cost_price=to_decimal(base_row.get("cost price")),
                        product_type=clean_str_with_strip(base_row.get("product_type")),
                        status=clean_str(base_row.get("product status")),
                        weight=to_decimal(base_row.get("package weight (kg)")),
                        length=to_decimal(base_row.get("package length(cms)")),
                        width=to_decimal(base_row.get("package width(cms)")),
                        height=to_decimal(base_row.get("package height(cms)")),
                        bundle_group_code=clean_str_with_strip(base_row.get("bundle / combo indicator")),
                        supplier=clean_str_with_strip(base_row.get("supplier name")),
                        country_of_origin=clean_str_with_strip(base_row.get("country of origin")),
                        product_id_type=clean_str_with_strip(base_row.get("product id type")),
                        product_condition=clean_str_with_strip(base_row.get("condition")),
                        hs_code=clean_str_with_strip(base_row.get("harmonized code (hs code)")),
                        category_id=category_id,
                        brand_id=brand_id,
                        stock=to_int(base_row.get("stock quantity")) or 0,
                        handling_time_days=to_int(base_row.get("handling time (days)")) or 0,
                        fast_dispatch=to_bool(base_row.get("fast dispatch")) or False,
                        free_shipping=to_bool(base_row.get("free shipping")) or False,
                        ships_from_location=clean_str_with_strip(base_row.get("ships from location")),
                        precautionary_note=clean_str_with_strip(base_row.get("precautionary note")),
                        care_instructions=clean_str_with_strip(get_first(base_row, ["care instructions (if applicable)", "care instructions (if any)"])),
                        warranty=clean_str_with_strip(get_first(base_row, ["warranty / guarantee (if applicable)", "warranty / guarantee"])),
                        product_code=clean_str_with_strip(base_row.get("product code")),
                        rrp_price=to_decimal(base_row.get("rrp")),
                        ean=None if any(clean_str_with_strip(r.get("parent sku / variation group code")) is not None for _, r in grouped_rows) else clean_str_with_strip(base_row.get("product code")),
                        tags=[clean_str_with_strip(base_row.get("product tag"))] if base_row.get("product tag") else None
                    )

                    db.add(product)
                    await db.flush()
                    products_map[product.sku] = product
                    created_products += 1
                    await log_activity(db, "product", product.id, "create", {"sku": product.sku, "title": product.title, "event": "bulk_import"}, performed_by=admin_id)

                # -----------------------------
                # SEO HANDLING
                # -----------------------------
                seo_keywords = clean_str_with_strip(base_row.get("seo keywords"))
                page_title = clean_str_with_strip(base_row.get("page title"))
                meta_description = clean_str_with_strip(base_row.get("meta description"))
                url_handle = clean_str_with_strip(base_row.get("url handles"))
                canonical_url = clean_str_with_strip(base_row.get("canonical url"))

                if any([seo_keywords, page_title, meta_description, url_handle, canonical_url]):
                    existing_seo = seo_map.get(product.id)
                    if existing_seo:
                            if seo_keywords is not None: existing_seo.meta_keywords = seo_keywords
                            if page_title is not None: existing_seo.page_title = page_title
                            if meta_description is not None: existing_seo.meta_description = meta_description
                            if url_handle is not None: existing_seo.url_handle = url_handle
                            if canonical_url is not None: existing_seo.canonical_url = canonical_url
                    else:
                        new_seo = ProductSEO(
                            product_id=product.id,
                            meta_keywords=seo_keywords,
                            page_title=page_title,
                            meta_description=meta_description,
                            url_handle=url_handle,
                            canonical_url=canonical_url
                        )
                        db.add(new_seo)
                        seo_map[product.id] = new_seo

                # -----------------------------
                # IMAGE HANDLING
                # -----------------------------
                images_to_add = []
                product_video_url = clean_str_with_strip(base_row.get("product video link"))
                if product_video_url:
                    images_to_add.append(ProductImage(product_id=product.id, image_url=None, video_url=product_video_url, is_main=True, image_order=0))

                order = 1
                for col in IMAGE_COLUMNS:
                    url_raw = base_row.get(col)
                    if url_raw:
                        url = clean_str_with_strip(url_raw)
                        if url:
                            # Use GCS URL if available
                            final_url = downloaded_images.get(url, url)
                            images_to_add.append(ProductImage(product_id=product.id, image_url=final_url, video_url=None, is_main=(col == "main image url"), image_order=order))
                            order += 1

                if images_to_add:
                    await db.execute(delete(ProductImage).where(ProductImage.product_id == product.id))
                    for img in images_to_add:
                        db.add(img)

                # -----------------------------
                # VARIANTS HANDLING
                # -----------------------------
                for v_no, v_row in grouped_rows:
                    v_sku = clean_str_with_strip(v_row.get("sku"))
                    if not v_sku or v_sku == product.sku:
                        continue

                    variant = variants_map.get(v_sku)
                    if variant:
                        v_title = clean_str_with_strip(v_row.get("title"))
                        if v_title: variant.title = v_title
                        v_sell = to_decimal(v_row.get("selling price"))
                        if v_sell: variant.price = v_sell
                        v_stock = to_int(v_row.get("stock quantity"))
                        if v_stock is not None: variant.stock = v_stock
                        v_ean = clean_str_with_strip(v_row.get("product code"))
                        if v_ean: variant.ean = v_ean
                        v_rrp = to_decimal(v_row.get("rrp"))
                        if v_rrp: variant.rrp_price = v_rrp
                        v_w = to_decimal(v_row.get("package weight (kg)"))
                        if v_w: variant.weight = v_w
                        v_l = to_decimal(v_row.get("package length(cms)"))
                        if v_l: variant.length = v_l
                        v_wi = to_decimal(v_row.get("package width(cms)"))
                        if v_wi: variant.width = v_wi
                        v_h = to_decimal(v_row.get("package height(cms)"))
                        if v_h: variant.height = v_h
                        updated_variants += 1
                    else:
                        variant = ProductVariant(
                            product_id=product.id,
                            sku=v_sku,
                            title=clean_str_with_strip(v_row.get("title")) or product.title,
                            price=to_decimal(v_row.get("selling price")) or 0,
                            stock=to_int(v_row.get("stock quantity")) or 0,
                            ean=clean_str_with_strip(v_row.get("product code")),
                            weight=to_decimal(v_row.get("package weight (kg)")),
                            length=to_decimal(v_row.get("package length(cms)")),
                            width=to_decimal(v_row.get("package width(cms)")),
                            height=to_decimal(v_row.get("package height(cms)")),
                            rrp_price=to_decimal(v_row.get("rrp")),
                        )
                        db.add(variant)
                        await db.flush()
                        variants_map[v_sku] = variant
                        created_variants += 1

                    # Variant Attributes - Support multiple header styles
                    await db.execute(delete(Attribute).where(Attribute.variant_id == variant.id))
                    for i in range(1, 4):
                        # Try Style 1: From export template (Variant Option X)
                        attr_name = clean_str_with_strip(v_row.get(f"variant option {i} (e.g., colour, size, style)"))
                        attr_val = clean_str_with_strip(v_row.get(f"variant values {i}"))
                        
                        # Try Style 2: Legacy style (attribute X name)
                        if not attr_name:
                            attr_name = clean_str_with_strip(v_row.get(f"attribute {i} name"))
                            attr_val = clean_str_with_strip(v_row.get(f"attribute {i} value"))

                        if attr_name and attr_val:
                            db.add(Attribute(name=attr_name, value=attr_val, variant_id=variant.id))

                    # Variant Images
                    variant_images_to_add = []
                    v_order = 1
                    for v_col in IMAGE_COLUMNS:
                        img_url_raw = v_row.get(v_col)
                        if img_url_raw:
                            img_url = clean_str_with_strip(img_url_raw)
                            if img_url:
                                # Use GCS URL if available
                                final_v_url = downloaded_images.get(img_url, img_url)
                                variant_images_to_add.append(ProductVariantImage(variant_id=variant.id, image_url=final_v_url, video_url=None, is_main=(v_col == "main image url"), image_order=v_order))
                                v_order += 1

                    if variant_images_to_add:
                        await db.execute(delete(ProductVariantImage).where(ProductVariantImage.variant_id == variant.id))
                        for v_img in variant_images_to_add:
                            db.add(v_img)

            except (IntegrityError, PendingRollbackError) as e:
                logger.error(f"Integrity/Rollback error processing product {product_key}: {e}")
                errors.append(f"Row {row_no}: Product SKU {product_key} failed database constraints and was skipped.")
                continue
            except SQLAlchemyError:
                raise 
            except Exception as e:
                logger.error(f"Error processing product {product_key}: {e}", exc_info=True)
                errors.append(f"Row {row_no}: Unexpected error for SKU {product_key}: {str(e)}")
                continue


    return {
        "created_products": created_products,
        "updated_products": updated_products,
        "created_variants": created_variants,
        "updated_variants": updated_variants,
        "deleted_products": deleted_products,
        "deleted_variants": deleted_variants,
        "errors": errors,
    }
async def _run_import(file_path: str, job_id: str, admin_id: str = "system"):
    engine, SessionLocal = create_sessionmaker()
    logger.info("job=%s started", job_id)

    loop = asyncio.get_running_loop()

    # 🔹 Read file OUTSIDE DB session
    rows = await loop.run_in_executor(None, _read_file_rows, file_path)
    total_rows = len(rows)

    # 🔹 Redis (sync → executor)
    await loop.run_in_executor(
        None, redis_client.hset, job_id, "total_rows", total_rows
    )

    async with SessionLocal() as db:
        await db.execute(
            update(BackgroundTask)
            .where(BackgroundTask.task_id == job_id)
            .values(
                status="RUNNING",
                task_info={"total_rows": total_rows, "processed_rows": 0}
            )
        )
        await db.commit()

        is_valid, errors, brand_map, category_map = await validate_brand_and_category(
            db, rows, Brand, Category
        )
        brand_map = brand_map or {}
        category_map = category_map or {}

        # 🔹 Pre-fetch all existing unique codes once for the entire import
        existing_codes_res = await db.execute(select(Product.unique_code).where(Product.unique_code != None))
        existing_codes = set(existing_codes_res.scalars().all())

        await db.commit()
        print("is_valid", is_valid)
        print("errors", errors)
#        print("brand_map", brand_map)

        # if not is_valid:
        #     '''
        #     await loop.run_in_executor(
        #         None,
        #         redis_client.hset,
        #         job_id,
        #         {"status": "failed", "errors": str(errors) if errors else ""},
        #     )'''
            # await loop.run_in_executor(
            #     None,
            #     lambda: redis_client.hset(
            #         job_id,
            #         mapping={
            #             "status": "failed",
            #             "errors": str(errors) if errors else "",
            #         },
            #     ),
            # )
            # await db.execute(
            #     update(BackgroundTask)
            #     .where(BackgroundTask.task_id == job_id)
            #     .values(status="FAILED", task_info=errors),
            # )
            # await db.commit()
            # return

    processed = 0
    BATCH_SIZE = 10
    NUM_WORKERS = 2
    worker_batches = [[] for _ in range(NUM_WORKERS)]

    for idx, row in enumerate(rows):
        sku_raw = row.get("sku")
        parent_code_raw = row.get("parent sku / variation group code")
        if parent_code_raw:
            group_key = clean_str_with_strip(parent_code_raw)
        else:
            group_key = clean_str_with_strip(sku_raw) if sku_raw else str(idx)
            
        worker_idx = hash(group_key) % NUM_WORKERS
        worker_batches[worker_idx].append((idx + 2, row))

    aggregated_result = {
        "created_products": 0,
        "updated_products": 0,
        "created_variants": 0,
        "updated_variants": 0,
        "deleted_products": 0,
        "deleted_variants": 0,
        "errors": []
    }

    last_db_update = 0
    update_lock = asyncio.Lock()

    PAUSE_POLL_INTERVAL = 2  # seconds

    async def worker_task(worker_rows):
        nonlocal processed, last_db_update
        for i in range(0, len(worker_rows), BATCH_SIZE):
            batch = worker_rows[i : i + BATCH_SIZE]

            # ── ⏸ Pause / 🛑 Abort control check ──────────────────────────
            while True:
                raw_ctrl = await loop.run_in_executor(
                    None, redis_client.get, f"{job_id}:ctrl"
                )
                ctrl = raw_ctrl.decode() if raw_ctrl else "running"

                if ctrl == "abort":
                    logger.info("job=%s aborted by user, stopping worker", job_id)
                    async with SessionLocal() as db_abort:
                        try:
                            await db_abort.execute(
                                update(BackgroundTask)
                                .where(BackgroundTask.task_id == job_id)
                                .values(
                                    status="ABORTED",
                                    task_info={
                                        "total_rows": total_rows,
                                        "processed_rows": processed,
                                        "errors": list(aggregated_result["errors"]),
                                        "created_products": aggregated_result["created_products"],
                                        "updated_products": aggregated_result["updated_products"],
                                        "created_variants": aggregated_result["created_variants"],
                                        "updated_variants": aggregated_result["updated_variants"],
                                        "deleted_products": aggregated_result["deleted_products"],
                                        "deleted_variants": aggregated_result["deleted_variants"],
                                    },
                                )
                            )
                            await db_abort.commit()
                        except Exception as db_e:
                            logger.error(f"Failed to mark task ABORTED in DB: {db_e}")
                    return  # exit this worker coroutine

                if ctrl != "pause":
                    break  # "running" — proceed with batch

                logger.info("job=%s paused, polling every %ss", job_id, PAUSE_POLL_INTERVAL)
                await asyncio.sleep(PAUSE_POLL_INTERVAL)
            # ── end pause/abort check ──────────────────────────────────────

            # 🔹 1. Collect all unique image URLs for this batch (OUTSIDE DB session)
            all_image_urls = set()
            for _, row in batch:
                for col in IMAGE_COLUMNS:
                    url_raw = row.get(col)
                    if url_raw:
                        url_str = clean_str_with_strip(url_raw)
                        if url_str:
                            all_image_urls.add(url_str)

            # 🔹 2. Download all images for this batch (OUTSIDE DB session)
            downloaded_images = {}
            if all_image_urls:
                async def _dl(u):
                    try:
                        gcs_u = await download_and_upload_image(u)
                        return u, gcs_u
                    except Exception as e:
                        logger.error(f"Image pre-fetch failed for {u}: {e}")
                        return u, u
                
                results = await asyncio.gather(*(_dl(u) for u in all_image_urls))
                downloaded_images = dict(results)

            max_retries = 3
            for attempt in range(max_retries):
                async with SessionLocal() as db:
                    try:
                        res = await process_batch(
                            db,
                            batch,
                            brand_map,
                            category_map,
                            existing_codes,
                            downloaded_images,
                            admin_id
                        )
                        for k, v in res.items():
                            if k == "errors":
                                aggregated_result["errors"].extend(v)
                            else:
                                aggregated_result[k] += v
                        break # Success
                    except Exception as e:
                        err_str = str(e)
                        if "1213" in err_str and attempt < max_retries - 1:
                            logger.warning(f"Deadlock detected in batch processing (attempt {attempt+1}). Retrying...")
                            await asyncio.sleep(0.5 * (attempt + 1))
                            continue
                        
                        logger.error(f"Batch processing failed after {attempt+1} attempts: {e}", exc_info=True)
                        aggregated_result["errors"].append(f"Batch failed: {str(e)}")
                        break
            
            processed += len(batch)
            await loop.run_in_executor(
                None, redis_client.hset, job_id, "processed_rows", processed
            )

            # Update BackgroundTask info in DB every 100 rows or at the end
            async with update_lock:
                if (processed - last_db_update) >= 100 or processed >= total_rows:
                    last_db_update = processed
                    async with SessionLocal() as db_update:
                        try:
                            await db_update.execute(
                                update(BackgroundTask)
                                .where(BackgroundTask.task_id == job_id)
                                .values(
                                    task_info={
                                        "total_rows": total_rows,
                                        "processed_rows": processed,
                                        "errors": list(aggregated_result["errors"]),
                                        "created_products": aggregated_result["created_products"],
                                        "updated_products": aggregated_result["updated_products"],
                                        "created_variants": aggregated_result["created_variants"],
                                        "updated_variants": aggregated_result["updated_variants"],
                                        "deleted_products": aggregated_result["deleted_products"],
                                        "deleted_variants": aggregated_result["deleted_variants"],
                                    },
                                )
                            )
                            await db_update.commit()
                        except Exception as db_e:
                            logger.error(f"@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@Failed to update task status in DB: {db_e}")

    tasks = [asyncio.create_task(worker_task(w_rows)) for w_rows in worker_batches if w_rows]
    await asyncio.gather(*tasks)

    async with SessionLocal() as db:
        await db.execute(
            update(BackgroundTask)
            .where(BackgroundTask.task_id == job_id)
            .values(
                status="COMPLETED",
                task_info={
                    "total_rows": total_rows,
                    "processed_rows": processed,
                    "errors": list(aggregated_result["errors"]),
                    "created_products": aggregated_result["created_products"],
                    "updated_products": aggregated_result["updated_products"],
                    "created_variants": aggregated_result["created_variants"],
                    "updated_variants": aggregated_result["updated_variants"],
                    "deleted_products": aggregated_result["deleted_products"],
                    "deleted_variants": aggregated_result["deleted_variants"],
                },
            )
        )
        await db.commit()

    await engine.dispose()


