# apis/v1/product_import_export.py
import io
import pandas as pd
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
import re
import unicodedata
from models.product import Product
from models.seo import ProductSEO
from models.b_tasks import BackgroundTask
from deps import get_db
from utils.admin_auth import require_catalog_supervisor
from utils.activity_logger import log_activity
from fastapi.responses import StreamingResponse
import csv
from datetime import datetime
from fastapi import status
from sqlalchemy.orm import selectinload
from models.product import ProductVariant
from models.category import Category
from models.brand import Brand
from typing import List, Dict, Any, Annotated, Optional, Literal, Tuple, Set
from io import StringIO, BytesIO
from openpyxl import load_workbook
from decimal import Decimal
from fastapi import APIRouter, Query, Depends, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from openpyxl import Workbook
from openpyxl.styles import Font
from schemas.import_export import (
    ProductImportRow,
    format_pydantic_error,
    BulkTemplateRequest,
    ProductExportRequest,
)
from utils.import_export import _read_file_rows
from models.product import ProductImage, ProductVariantImage
from utils.image_handler import download_and_upload_image
from urllib.parse import urlparse
from sqlalchemy import func
from celery_worker.celery_app import export_products_task
from service.product_import_export import build_export_data, write_file_for_response
from apis.v1.utils import generate_unique_product_code


def is_valid_url(url: str) -> bool:
    try:
        parsed = urlparse(url.strip())
        return parsed.scheme in ("http", "https") and bool(parsed.netloc)
    except Exception:
        return False


router = APIRouter()


async def validate_brand_and_category(db, rows, Brand, Category):
    brand_names: Set[str] = set()
    category_codes: Set[str] = set()

    for r in rows:
        b = r.get("brand name")
        c = r.get("category code")

        if b:
            brand_names.add(str(b).strip().lower())
        if c:
            category_codes.add(str(c).strip().lower())

    errors = []

    brand_map = {}
    category_map = {}

    # -------- BRANDS --------
    if brand_names:
        q = await db.execute(
            select(Brand.id, Brand.name).where(func.lower(Brand.name).in_(brand_names))
        )

        rows_found = q.all()
        brand_map = {name.lower(): bid for bid, name in rows_found}

        missing = brand_names - set(brand_map.keys())
        for m in missing:
            errors.append(f"Brand name '{m}' does not exist in database.")

    # -------- CATEGORIES --------
    if category_codes:
        q = await db.execute(
            select(Category.id, Category.category_code).where(
                func.lower(func.trim(Category.category_code)).in_(category_codes)
            )
        )

        rows_found = q.all()
        # Ensure we strip() the database code when building the map, as spaces might exist
        category_map = {c_code.strip().lower(): cid for cid, c_code in rows_found if c_code}
        missing = category_codes - set(category_map.keys())
        for m in missing:
            errors.append(f"Category code '{m}' does not exist in database.")

    if errors:
        return False, errors, None, None

    return True, [], brand_map, category_map


def make_slug(value: str | None) -> str:
    if not value:
        return ""
    value = unicodedata.normalize("NFKD", value)
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = value.strip("-")
    value = re.sub(r"-+", "-", value)
    return value


def clean_str_with_strip(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
    else:
        s = str(value).strip()
    return s if s else None


def clean_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
    else:
        s = str(value).strip()
    return s.lower() if s else None


def to_int(value: Any) -> int | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(float(s))  # handles "10.0" too
    except ValueError:
        return None


def to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def to_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if not s:
        return None
    if s in ("yes", "y", "true", "1"):
        return True
    if s in ("no", "n", "false", "0"):
        return False
    return None


@router.post(
    "/import-products",
    status_code=status.HTTP_201_CREATED,
    responses={
        400: {"description": "File is empty or could not be parsed."},
    },
)
async def import_products(
    file: Annotated[UploadFile, File(...)],
    db: Annotated[AsyncSession, Depends(get_db)],
    user: Annotated[dict, Depends(require_catalog_supervisor)],
):
    rows = await _read_file_rows(file)

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File is empty or could not be parsed.",
        )

    errors: List[str] = []
    created_products = 0
    created_variants = 0
    for idx, row in enumerate(rows):
        try:
            validated = ProductImportRow(**row)
        except Exception as e:
            simple_errors = format_pydantic_error(e)
            for m in simple_errors:
                errors.append(f"Row {idx + 2}: {m}")
            continue
    if errors:
        return {
            "created_products": created_products,
            "created_variants": created_variants,
            "errors": errors,
        }
    is_valid, validation_errors, brand_map, category_map = (
        await validate_brand_and_category(db, rows, Brand, Category)
    )

    if not is_valid:
        # append these errors to your main error list
        errors.extend(validation_errors)

        # return exactly the same format your frontend expects
        return {
            "created_products": created_products,
            "created_variants": created_variants,
            "errors": errors,
        }
    await db.commit()
    # -------- group rows: same variation group => same product (variants) -------
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for idx, row in enumerate(rows):
        sku = row.get("sku")
        parent_code = row.get("parent sku / variation group code")

        if parent_code:
            group_key = parent_code  # ONLY variation code used
        else:
            # if no variation code → generate unique group key for this row
            group_key = f"NO_VARIATION_{idx}_{sku}"

        grouped.setdefault(group_key, []).append(row)

    # -------- process grouped rows: create products and variants -------

    image_columns = [
        "main image url",
        "lifestyle image url",
        "image 1 url",
        "image 2 url",
        "image 3 url",
        "image 4 url",
        "image 5 url",
        "image 6 url",
        "image 7 url",
        "image 8 url",
        "image 9 url",
        "image 10 url",
    ]
    # -----------------------------
    # PRE-FETCH EXISTING UNIQUE CODES
    # -----------------------------
    stmt = select(Product.unique_code).where(Product.unique_code != None)
    result = await db.execute(stmt)
    existing_codes = set(result.scalars().all())

    async with db.begin():
        # product_sku here is actually the group key (variation group code)
        for product_sku, sku_rows in grouped.items():
            base_row = sku_rows[0]
            product_video_url = clean_str_with_strip(base_row.get("product video link"))
            # ---------------- PRODUCT ----------------
            # check product SKU uniqueness
            product_id = clean_str_with_strip(base_row.get("product code"))
            # brand_id = clean_str_with_strip(base_row.get("brand name"))
            # category_code = clean_str_with_strip(base_row.get("category code"))
            existing_product = await db.execute(
                select(Product).where(Product.sku == product_sku)
            )
            if existing_product.scalar_one_or_none():
                errors.append(f"Product SKU '{product_sku}' already exists. Skipping.")
                continue
            existing_product = await db.execute(
                select(Product).where(Product.id == product_id)
            )
            if existing_product.scalar_one_or_none():
                errors.append(f"Product ID '{product_id}' already exists. Skipping.")
                continue
            # generate unique code
            unique_code = await generate_unique_product_code(db)
            new_product = Product(
                unique_code=unique_code,
                sku=clean_str_with_strip(base_row.get("sku")),
                variation_group_code=(
                    clean_str_with_strip(parent_code) if parent_code else None
                ),
                product_code=clean_str_with_strip(base_row.get("product code")),
                ean=clean_str_with_strip(base_row.get("product code")),
                title=clean_str_with_strip(base_row.get("title")),
                slug=make_slug(clean_str(base_row.get("title"))),
                description=clean_str_with_strip(base_row.get("long description")),
                price=to_decimal(base_row.get("selling price")),
                product_type=clean_str_with_strip(base_row.get("product_type")),
                cost_price=to_decimal(base_row.get("cost price")),
                rrp_price=to_decimal(base_row.get("rrp")),
                status=clean_str(base_row.get("product status")),
                weight=to_decimal(base_row.get("package weight (kg)")),
                length=to_decimal(base_row.get("package length(cms)")),
                width=to_decimal(base_row.get("package width(cms)")),
                height=to_decimal(base_row.get("package height(cms)")),
                bundle_group_code=clean_str_with_strip(
                    base_row.get("bundle / combo indicator")
                ),
                unit="kg",
                supplier=clean_str_with_strip(base_row.get("supplier name")),
                country_of_origin=clean_str_with_strip(
                    base_row.get("country of origin")
                ),
                product_id_type=clean_str_with_strip(base_row.get("product id type")),
                asin=None,
                mpn=None,
                hs_code=clean_str_with_strip(base_row.get("harmonized code (hs code)")),
                category_id=(category_map or {}).get(clean_str(base_row.get("category code"))),
                brand_id=(brand_map or {}).get(clean_str(base_row.get("brand name"))),
                stock=to_int(base_row.get("stock quantity")),
                key_features=None,
                is_battery_required=to_bool(base_row.get("is battery required?")),
                shipping_template=clean_str_with_strip(
                    base_row.get("shipping template")
                ),
                precautionary_note=clean_str(base_row.get("precautionary note")),
                care_instructions=clean_str(
                    base_row.get("care instructions (if applicable)")
                ),
                warranty=clean_str(
                    base_row.get("warranty / guarantee (if applicable)")
                ),
                # if you want to store Product Tag as JSON list:
                tags=(
                    [clean_str_with_strip(base_row.get("product tag"))]
                    if base_row.get("product tag")
                    else None
                ),
                ships_from_location=clean_str_with_strip(
                    base_row.get("ships from location")
                ),
                handling_time_days=to_int(base_row.get("handling time (days)")),
                fast_dispatch=to_bool(base_row.get("fast dispatch")),
                free_shipping=to_bool(base_row.get("free shipping")),
                product_condition=clean_str_with_strip(base_row.get("condition")),
                estimated_shipping_cost=None,
                product_margin_percent=None,
                product_margin_amount=None,
                profit=None,
            )
            db.add(new_product)
            await db.flush()  # new_product.id
            await db.refresh(new_product)
            await log_activity(db, "product", new_product.id, "create", {"sku": new_product.sku, "title": new_product.title, "event": "bulk_import_sync"}, performed_by=user.get("admin_id"))

            # ---------------- SEO HANDLING ----------------
            seo_keywords = clean_str_with_strip(base_row.get("seo keywords"))
            page_title = clean_str_with_strip(base_row.get("page title"))
            meta_description = clean_str_with_strip(base_row.get("meta description"))
            url_handle = clean_str_with_strip(base_row.get("url handles"))
            canonical_url = clean_str_with_strip(base_row.get("canonical url"))

            if any([seo_keywords, page_title, meta_description, url_handle, canonical_url]):
                db.add(
                    ProductSEO(
                        product_id=new_product.id,
                        meta_keywords=seo_keywords,
                        page_title=page_title,
                        meta_description=meta_description,
                        url_handle=url_handle,
                        canonical_url=canonical_url,
                    )
                )

            created_products += 1

            # ---------------- PRODUCT-LEVEL IMAGES ----------------
            order = 1
            added_any = False

            for col in image_columns:
                url = clean_str_with_strip(base_row.get(col))

                is_main = col == "main image url"
                video = product_video_url if is_main else None

                # Skip only when BOTH image and video are missing
                if not url and not video:
                    continue

                db.add(
                    ProductImage(
                        product_id=new_product.id,
                        image_url=url,  # allowed to be None
                        is_main=is_main,
                        video_url=video,
                        image_order=order,
                    )
                )

                order += 1
                added_any = True

            # Fallback: video-only product (no images at all)
            if not added_any and product_video_url:
                db.add(
                    ProductImage(
                        product_id=new_product.id,
                        image_url=None,
                        is_main=True,
                        video_url=product_video_url,
                        image_order=1,
                    )
                )

            # ---------------- VARIANTS + ATTRIBUTES + VARIANT IMAGES -------------------
            for row in sku_rows:
                variant_sku = clean_str_with_strip(row.get("sku"))
                if not variant_sku:
                    continue

                # check variant SKU uniqueness
                existing_variant = await db.execute(
                    select(ProductVariant).where(ProductVariant.sku == variant_sku)
                )
                if existing_variant.scalar_one_or_none():
                    errors.append(
                        f"Variant SKU '{variant_sku}' already exists. Skipping."
                    )
                    continue

                new_variant = ProductVariant(
                    product_id=new_product.id,
                    title=clean_str_with_strip(row.get("title")),
                    sku=variant_sku,
                    ean=clean_str_with_strip(row.get("product code")),
                    price=to_decimal(row.get("selling price")),
                    # rrp_price=to_decimal(row.get("rrp")),
                    stock=to_int(
                        row.get("stock quantity") or base_row.get("stock quantity")
                    ),
                    cost_price=0,
                    ships_from_location=clean_str_with_strip(
                        row.get("ships from location")
                        or base_row.get("ships from location")
                    ),
                    handling_time_days=to_int(
                        row.get("handling time (days)")
                        or base_row.get("handling time (days)")
                    ),
                    width=to_decimal(row.get("package width(cms)")),
                    height=to_decimal(row.get("package height(cms)")),
                    weight=to_decimal(row.get("package weight (kg)")),
                    length=to_decimal(row.get("package length(cms)")),
                    bundle_group_code=clean_str_with_strip(
                        base_row.get("bundle / combo indicator")
                    ),
                    estimated_shipping_cost=None,
                    product_margin_percent=None,
                    product_margin_amount=None,
                    profit=None,
                    rrp_price=to_decimal(row.get("rrp") or base_row.get("rrp")),
                    free_shipping=to_bool(base_row.get("free shipping")),
                    fast_dispatch=to_bool(base_row.get("fast dispatch")),
                    hs_code=clean_str_with_strip(
                        base_row.get("harmonized code (hs code)")
                    ),
                )
                db.add(new_variant)
                await db.flush()
                created_variants += 1

                # ---------- ATTRIBUTES ----------
                max_options = 10  # supports up to 10 options; increase if needed

                for i in range(1, max_options + 1):
                    option_col = f"variant option {i} (e.g., colour, size, style)"
                    value_col = f"variant values {i}"

                    name = clean_str_with_strip(row.get(option_col))
                    value = clean_str_with_strip(row.get(value_col))

                    if not name or not value:
                        continue

                    db.add(
                        Attribute(
                            name=name,  # e.g. "colour", "size"
                            value=value,  # e.g. "black", "xl"
                            variant_id=new_variant.id,
                        )
                    )

                # ---------- VARIANT-LEVEL IMAGES ----------
                order = 1
                added_any = False

                for col in image_columns:
                    url = clean_str_with_strip(base_row.get(col))

                    is_main = col == "main image url"
                    video = product_video_url if is_main else None

                    # Skip only when BOTH image and video are missing
                    if not url and not video:
                        continue

                    db.add(
                        ProductVariantImage(
                            variant_id=new_variant.id,
                            image_url=url,  # allowed to be None
                            is_main=is_main,
                            video_url=video,
                            image_order=order,
                        )
                    )

                    order += 1
                    added_any = True

                # Fallback: video-only product (no images at all)
                if not added_any and product_video_url:
                    db.add(
                        ProductImage(
                            product_id=new_product.id,
                            image_url=None,
                            is_main=True,
                            video_url=product_video_url,
                            image_order=1,
                        )
                    )

    return {
        "created_products": created_products,
        "created_variants": created_variants,
        "errors": errors,
    }


# @router.get("/export-products")
# async def export_products(
#     # /export-products?product_ids=uuid1&product_ids=uuid2
#     product_ids: Optional[List[str]] = Query(
#         None,
#         description="List of product IDs to export. If empty, exports all products.",
#     ),
#     download_flag: Literal["csv", "excel"] = Query(
#         "excel",
#         description="File type to download: 'csv' or 'excel'. Default is 'excel'.",
#     ),
#     db: AsyncSession = Depends(get_db),
# ):
#     """
#     Export products (optionally filtered by IDs) with their variants & attributes.
#     First 3 columns: category_code, category(breadcrumb), product_type.
#     File type controlled by `download_flag` (default: excel).
#     """

#     # --- Query with eager loading ---
#     query = select(Product).options(
#         selectinload(Product.category),
#         selectinload(Product.images),
#         selectinload(Product.variants).selectinload(ProductVariant.attributes),
#         selectinload(Product.variants).selectinload(ProductVariant.images),
#     )

#     if product_ids:
#         query = query.where(Product.id.in_(product_ids))

#     result = await db.execute(query)
#     products: List[Product] = result.scalars().all()

#     rows: List[List[str]] = []

#     # Small helpers
#     def _bool_to_flag(val: Optional[bool]) -> str:
#         if val is None:
#             return ""
#         return "Yes" if val else "No"

#     def _dec_or_empty(val) -> str:
#         return "" if val is None else str(val)

#     def _join_tags(tags) -> str:
#         if not tags:
#             return ""
#         if isinstance(tags, list):
#             return ", ".join(str(t) for t in tags)
#         return str(tags)

#     def _collect_image_urls(
#         product: Product, variant: Optional[ProductVariant]
#     ) -> List[str]:
#         """
#         Returns list of image URLs in priority order:
#         1. Variant main image, then other variant images
#         2. Product main image, then other product images
#         """
#         urls: List[str] = []

#         # Variant images first
#         if variant and variant.images:
#             main = [
#                 img.image_url for img in variant.images if img.is_main and img.image_url
#             ]
#             others = [
#                 img.image_url
#                 for img in variant.images
#                 if not img.is_main and img.image_url
#             ]
#             urls.extend(main + others)

#         # Product images next
#         if product.images:
#             main = [
#                 img.image_url for img in product.images if img.is_main and img.image_url
#             ]
#             others = [
#                 img.image_url
#                 for img in product.images
#                 if not img.is_main and img.image_url
#             ]
#             urls.extend(main + others)

#         # Deduplicate while preserving order
#         seen = set()
#         ordered = []
#         for u in urls:
#             if u not in seen:
#                 seen.add(u)
#                 ordered.append(u)
#         return ordered

#     def _collect_video_url(product: Product, variant: Optional[ProductVariant]) -> str:
#         # Variant video_url first
#         if variant and variant.images:
#             for img in variant.images:
#                 if img.video_url:
#                     return img.video_url
#         # then product
#         if product.images:
#             for img in product.images:
#                 if img.video_url:
#                     return img.video_url
#         return ""

#     # --- Build rows ---
# for product in products:
#     # Category info (first 3 columns)
#     if product.category_id:
#         category_code, category_breadcrumb, leaf_product_type = (
#             await build_category_info(db, product.category_id)
#         )
#     else:
#         category_code, category_breadcrumb, leaf_product_type = "", "", ""

#     # base fields used for both product & variants
#     product_brand_code = (
#         product.brand_id or ""
#     )  # change if Brand has separate 'code'
#     product_code = product.product_code or ""
#     supplier_name = product.supplier or ""
#     product_tag = _join_tags(product.tags)
#     hs_code = product.hs_code or ""
#     shipping_template = product.shipping_template or ""
#     precautionary_note = product.precautionary_note or ""
#     care_instructions = product.care_instructions or ""
#     warranty = product.warranty or ""
#     # You can fetch SEO keywords from product.seo if defined
#     seo_keywords = ""  # placeholder
#     cost_price = product.cost_price or ""

#     # Get all images & video at product level (fallback)
#     product_level_imgs = _collect_image_urls(product, None)
#     product_level_video = _collect_video_url(product, None)

#     if product.variants:
#         for variant in product.variants:
#             # Attributes -> up to 2 options
#             attrs = list(variant.attributes or [])
#             attr1_name = attrs[0].name if len(attrs) >= 1 else ""
#             attr1_val = attrs[0].value if len(attrs) >= 1 else ""
#             attr2_name = attrs[1].name if len(attrs) >= 2 else ""
#             attr2_val = attrs[1].value if len(attrs) >= 2 else ""

#             # Price / RRP / stock / dimensions / shipping
#             selling_price = variant.price or product.price
#             rrp = variant.rrp_price or product.rrp_price
#             cost_price = variant.cost_price or product.cost_price
#             stock_qty = (
#                 variant.stock if variant.stock is not None else product.stock
#             )

#             weight = (
#                 variant.weight
#                 if variant.weight not in (None, 0)
#                 else product.weight
#             )
#             length = (
#                 variant.length
#                 if variant.length not in (None, 0)
#                 else product.length
#             )
#             width = (
#                 variant.width if variant.width not in (None, 0) else product.width
#             )
#             height = (
#                 variant.height
#                 if variant.height not in (None, 0)
#                 else product.height
#             )

#             handling_time = (
#                 variant.handling_time_days
#                 if variant.handling_time_days is not None
#                 else product.handling_time_days
#             )

#             ships_from_location = (
#                 variant.ships_from_location
#                 if variant.ships_from_location
#                 else product.ships_from_location or ""
#             )

#             # Images & video (variant overrides)
#             img_urls = _collect_image_urls(product, variant)
#             video_url = _collect_video_url(product, variant) or product_level_video

#             main_image = img_urls[0] if len(img_urls) >= 1 else ""
#             lifestyle_image = img_urls[1] if len(img_urls) >= 2 else ""

#             extra_imgs = img_urls[2:12]  # up to 10 images total
#             extra_imgs += [""] * (10 - len(extra_imgs))  # pad to 10

#             # Build row in EXACT header order
#             row = [
#                 # 1–3: category info
#                 category_code or "",  # "category code"
#                 category_breadcrumb or "",  # "category"
#                 leaf_product_type or "",  # "product type"
#                 # 4. "brand code"
#                 # product_brand_code,
#                 # 5. "SKU"
#                 variant.sku or product.sku or "",
#                 # 6. "Product Code"
#                 variant.ean or "",
#                 # 7. "Title"
#                 product.title or "",
#                 # 8. "Long Description"
#                 product.description or "",
#                 # 9. "Brand_id"
#                 product.brand_id or "",
#                 # 10. "Condition"
#                 product.product_condition or "",
#                 # 11. "Product Status"
#                 product.status or "",
#                 # 12. "Supplier Name"
#                 supplier_name,
#                 # 13. "Product Tag"
#                 product_tag,
#                 # 14. "Parent SKU / Variation Group Code"
#                 product.variation_group_code or "",
#                 # 15. "Variant Option 1"
#                 attr1_name,
#                 # 16. "Variant Values 1"
#                 attr1_val,
#                 # 17. "Variant Option 2"
#                 attr2_name,
#                 # 18. "Variant Values 2"
#                 attr2_val,
#                 # 19. "Bundle / Combo Indicator"
#                 "",
#                 # 20. "Selling Price"
#                 _dec_or_empty(selling_price),
#                 # 21. "RRP"
#                 _dec_or_empty(rrp),
#                 # 22. "Cost Price"
#                 _dec_or_empty(cost_price),
#                 # 23. "Stock Quantity"
#                 str(stock_qty or ""),
#                 # 24. "Package Weight (kg)"
#                 _dec_or_empty(weight),
#                 # 25. "Package Length(Cms)"
#                 _dec_or_empty(length),
#                 # 25. "Package Width(Cms)"
#                 _dec_or_empty(width),
#                 # 26. "Package Height(Cms)"
#                 _dec_or_empty(height),
#                 # 27. "Harmonized Code (HS Code)"
#                 hs_code,
#                 # 28. "Shipping Template"
#                 shipping_template,
#                 # 29. "Handling Time (days)"
#                 str(handling_time or "") if handling_time is not None else "",
#                 # 30. "Fast Dispatch"
#                 product.fast_dispatch if product.fast_dispatch is not None else "",
#                 # 31. "Free Shipping"
#                 product.free_shipping if product.free_shipping is not None else "",
#                 # 32. "Ships From Location"
#                 ships_from_location,
#                 # 33. "Product Video Link"
#                 video_url,
#                 # 34. "Main Image URL"
#                 main_image,
#                 # 35. "Lifestyle Image URL"
#                 lifestyle_image,
#                 # 36–45. "Image 1 URL"..."Image 10 URL"
#                 *extra_imgs,
#                 # 46. "Precautionary Note"
#                 precautionary_note,
#                 # 46. "Care Instructions (If Applicable)"
#                 care_instructions,
#                 # 47. "Warranty / Guarantee (If Applicable)"
#                 warranty,
#                 # 48. "SEO Keywords"
#                 seo_keywords,
#             ]

#             rows.append(row)
#     else:
#         # Product without variants -> one row representing base product
#         selling_price = product.price
#         rrp = product.rrp_price
#         cost_price = product.cost_price
#         stock_qty = product.stock

#         weight = product.weight
#         length = product.length
#         width = product.width
#         height = product.height
#         handling_time = product.handling_time_days
#         ships_from_location = product.ships_from_location or ""

#         img_urls = product_level_imgs
#         video_url = product_level_video

#         main_image = img_urls[0] if len(img_urls) >= 1 else ""
#         lifestyle_image = img_urls[1] if len(img_urls) >= 2 else ""
#         extra_imgs = img_urls[2:12]
#         extra_imgs += [""] * (10 - len(extra_imgs))

#         row = [
#             # 1–3: category info
#             category_code or "",
#             category_breadcrumb or "",
#             leaf_product_type or "",
#             # 5. "SKU"
#             product.sku or "",
#             # 6. "Product Code"
#             product_code,
#             # 7. "Title"
#             product.title or "",
#             # 8. "Long Description"
#             product.description or "",
#             # 9. "Brand_id"
#             product.brand_id or "",
#             # 10. "Condition"
#             product.product_condition or "",
#             # 11. "Product Status"
#             product.status or "",
#             # 12. "Supplier Name"
#             supplier_name,
#             # 13. "Product Tag"
#             product_tag,
#             # 14. "Parent SKU / Variation Group Code"
#             product.variation_group_code or product.sku or "",
#             # 15–18: no variant attributes
#             "",
#             "",
#             "",
#             "",
#             # 19. "Bundle / Combo Indicator"
#             "",
#             # 20. "Selling Price"
#             _dec_or_empty(selling_price),
#             # 21. "RRP"
#             _dec_or_empty(rrp),
#             _dec_or_empty(cost_price),
#             # 22. "Stock Quantity"
#             str(stock_qty or ""),
#             # 23–26: package dimensions
#             _dec_or_empty(weight),
#             _dec_or_empty(length),
#             _dec_or_empty(width),
#             _dec_or_empty(height),
#             # 27. "Harmonized Code (HS Code)"
#             hs_code,
#             # 28. "Shipping Template"
#             shipping_template,
#             # 29. "Handling Time (days)"
#             str(handling_time or "") if handling_time is not None else "",
#             # 30. "Fast Dispatch"
#             product.fast_dispatch if product.fast_dispatch is not None else "",
#             # 31. "Free Shipping"
#             product.free_shipping if product.free_shipping is not None else "",
#             # 32. "Ships From Location"
#             ships_from_location,
#             # 33. "Product Video Link"
#             video_url,
#             # 34. "Main Image URL"
#             main_image,
#             # 35. "Lifestyle Image URL"
#             lifestyle_image,
#             # 36–45: images
#             *extra_imgs,
#             # 46. "Precautionary Note"
#             precautionary_note,
#             # 46. "Care Instructions (If Applicable)"
#             care_instructions,
#             # 47. "Warranty / Guarantee (If Applicable)"
#             warranty,
#             # 48. "SEO Keywords"
#             seo_keywords,
#         ]

#         rows.append(row)

#     # --- Build response: CSV or Excel ---
#     timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#     CATEGORY_HEADERS = [
#         "Category code",
#         "category",
#     ]
#     new_headers = EXPECTED_HEADERS.copy()
#     new_headers[0] = "Product Type"
#     new_headers.pop(1)
#     export_headers = CATEGORY_HEADERS + new_headers

#     if rows and len(export_headers) != len(rows[0]):
#         raise HTTPException(
#             status_code=500,
#             detail=(
#                 f"Export schema mismatch: "
#                 f"{len(export_headers)} headers vs {len(rows[0])} columns"
#             ),
#         )
#     if download_flag == "csv":
#         output = io.StringIO()
#         writer = csv.writer(output)
#         writer.writerow(export_headers)
#         for row in rows:
#             writer.writerow(row)
#         output.seek(0)

#         filename = f"products_export_{timestamp}.csv"
#         return StreamingResponse(
#             output,
#             media_type="text/csv",
#             headers={"Content-Disposition": f'attachment; filename="{filename}"'},
#         )

#     # Excel (default)
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Products"

#     ws.append(export_headers)
#     bold_font = Font(bold=True)
#     for cell in ws[1]:  # first row = header
#         cell.font = bold_font
#     for row in rows:
#         ws.append(row)

#     stream = io.BytesIO()
#     wb.save(stream)
#     stream.seek(0)

#     filename = f"products_export_{timestamp}.xlsx"
#     return StreamingResponse(
#         stream,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={"Content-Disposition": f'attachment; filename="{filename}"'},
#     )


EXPECTED_HEADERS = [
    "category code",
    "brand code",
    "SKU",
    "Product Code",
    "Title",
    "Long Description",
    "Brand Name",
    "Condition",
    "Product Status",
    "Supplier Name",
    "Product Tag",
    "Parent SKU / Variation Group Code",
    "Variant Option 1 (e.g., Colour, Size, Style)",
    "Variant Values 1",
    "Variant Option 2 (e.g., Colour, Size, Style)",
    "Variant Values 2",
    "Bundle / Combo Indicator",
    "Selling Price",
    "RRP",
    "Cost Price",
    "Stock Quantity",
    "Package Weight (kg)",
    "Package Length(Cms)",
    "Package Width(Cms)",
    "Package Height(Cms)",
    "Harmonized Code (HS Code)",
    "Shipping Template",
    "Handling Time (days)",
    "Fast Dispatch",
    "Free Shipping",
    "Ships From Location",
    "Product Video Link",
    "Main Image URL",
    "Lifestyle Image URL",
    "Image 1 URL",
    "Image 2 URL",
    "Image 3 URL",
    "Image 4 URL",
    "Image 5 URL",
    "Image 6 URL",
    "Image 7 URL",
    "Image 8 URL",
    "Image 9 URL",
    "Image 10 URL",
    "Precautionary Note",
    "Care Instructions (If Applicable)",
    "Warranty / Guarantee (If Applicable)",
    "SEO Keywords",
    "Page Title",
    "Meta Description",
    "URL Handles",
    "Canonical URL",
]

# Drop category_id & brand_id from the "body" of the template
BASE_PRODUCT_HEADERS = [
    h for h in EXPECTED_HEADERS if h not in {"category code", "brand code"}
]

# New final headers for the download template (first 3 are new)
TEMPLATE_HEADERS = [
    "category code",  # leaf category id
    "category",  # breadcrumb: Parent > Child > Leaf
    "product_type",  # leaf category name
    *BASE_PRODUCT_HEADERS,
]


def build_category_info_from_map(
    category_id: str,
    category_map: dict,
) -> tuple[str, str, str]:
    """
    Pure in-memory version of build_category_info.
    Requires a pre-built {id: Category} map (no DB queries).

    Returns (category_identifier, breadcrumb, product_type)
    - category_identifier : category_code if set, otherwise the ID
    - breadcrumb          : "Root > Parent > ... > Leaf"
    - product_type        : leaf category name
    """
    category = category_map.get(category_id)
    if not category:
        return category_id, "", ""

    names: list[str] = []
    current = category
    seen: set[str] = set()   # cycle guard

    while current and current.id not in seen:
        seen.add(current.id)
        names.append(current.name)
        if current.parent_id is None:
            break
        current = category_map.get(current.parent_id)

    names.reverse()  # leaf→root collected, flip to root→leaf

    breadcrumb = " > ".join(names)
    product_type = category.name
    category_identifier = category.category_code or str(category.id)

    return category_identifier, breadcrumb, product_type


async def build_category_info(
    db: AsyncSession, category_id: str
) -> tuple[str, str, str]:
    """
    Convenience async wrapper (used by callers that don't have a map yet).
    Loads all categories in one query when called standalone.
    Prefer build_category_info_from_map when processing multiple IDs.
    """
    all_result = await db.execute(select(Category))
    category_map = {c.id: c for c in all_result.scalars().all()}
    return build_category_info_from_map(category_id, category_map)


@router.get("/download-products-template")
async def download_products_template(
    db: Annotated[AsyncSession, Depends(get_db)],
    file_type: Annotated[Literal["csv", "excel"], Query()] = "csv",
    category_ids: Annotated[
        list[str] | None,
        Query(description="Optional list of leaf category IDs to pre-fill template rows"),
    ] = None,
):

    # Build rows
    rows: list[list[str]] = []

    if category_ids:
        # Load all categories once — avoids N×depth DB queries in the loop
        all_cats_result = await db.execute(select(Category))
        category_map = {c.id: c for c in all_cats_result.scalars().all()}

        for cid in category_ids:
            cat_id, breadcrumb, product_type = build_category_info_from_map(cid, category_map)
            # 3 prefilled columns, rest empty
            rows.append(
                [cat_id, breadcrumb, product_type]
                + [""] * len(BASE_PRODUCT_HEADERS)
                + [""]  # brand_id (optional) at the end
            )
    else:
        # A single empty row
        rows.append(
            ["", "", ""]  # category_id, category, product_type
            + [""] * len(BASE_PRODUCT_HEADERS)
            + [""]
        )
    category_name = None
    if category_ids:
        # Reuse the map already built above (or fetch once if no category_ids processed)
        first_cat = category_map.get(category_ids[0]) if category_ids else None
        category_name = first_cat.name if first_cat else None
    filename = "products_template"
    if category_name:
        safe_name = category_name.lower().replace(" ", "_")
        filename = f"{safe_name}_template"
    if file_type == "csv":
        # ---- CSV (default) ----
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(TEMPLATE_HEADERS)
        writer.writerows(rows)
        output.seek(0)

        # Use iter([output.getvalue()]) so StreamingResponse can stream the data
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )

    # ---- Excel (.xlsx) ----
    output = io.BytesIO()
    wb = Workbook()

    # --------------------------
    # Sheet 1 → Products Template
    # --------------------------
    ws1 = wb.active
    ws1.title = "Products Template"

    # Headers
    ws1.append(TEMPLATE_HEADERS)
    # Bold header
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    # Rows
    for row in rows:
        ws1.append(row)

    # --------------------------
    # Sheet 2 → Brands List
    # --------------------------
    ws2 = wb.create_sheet("Brands")

    # Add brand headers
    ws2.append(["brand_id", "brand_name"])

    # Fetch all brands
    brand_rows = await db.execute(select(Brand))
    brands = brand_rows.scalars().all()

    # Add all brands
    for b in brands:
        ws2.append([str(b.id), b.name])

    # --------------------------
    # Sheet 3 → Categories List
    # --------------------------
    ws3 = wb.create_sheet("Categories")

    # Add category headers
    ws3.append(["category_id", "category_code", "category_name", "breadcrumb"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    # Fetch all categories to build breadcrumbs in memory
    all_category_rows = await db.execute(select(Category))
    all_categories_db = all_category_rows.scalars().all()
    category_map = {cat.id: cat for cat in all_categories_db}

    # Filter active and sort
    active_categories = sorted([c for c in all_categories_db if c.is_active], key=lambda x: x.name)

    # Add all categories with breadcrumbs
    for cat in active_categories:
        names = []
        curr = cat
        depth = 0
        while curr and depth < 20:
            names.append(curr.name)
            if curr.parent_id is None:
                break
            curr = category_map.get(curr.parent_id)
            depth += 1
        
        names.reverse()
        breadcrumb = " > ".join(names)
        cat_code = cat.category_code or str(cat.id)
        
        ws3.append([str(cat.id), cat_code, cat.name, breadcrumb])

    # --------------------------
    # Data Validation (Dropdowns)
    # --------------------------
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # 1. Category Code (Column A in Products Template)
    if len(active_categories) > 0:
        cat_code_range = f"'Categories'!$B$2:$B${len(active_categories) + 1}"
        cv = DataValidation(type="list", formula1=cat_code_range, allow_blank=True)
        cv.errorTitle = 'Invalid Category Code'
        cv.errorMessage = 'Please select a category code from the list.'
        ws1.add_data_validation(cv)
        # Apply to column A, rows 2 to 1000
        cv.add('A2:A1000')

    # 2. Brand Name (Column H in Products Template)
    if len(brands) > 0:
        brand_range = f"'Brands'!$B$2:$B${len(brands) + 1}"
        bv = DataValidation(type="list", formula1=brand_range, allow_blank=True)
        bv.errorTitle = 'Invalid Brand Name'
        bv.errorMessage = 'Please select a brand name from the list.'
        ws1.add_data_validation(bv)
        # Apply to column H, rows 2 to 1000
        bv.add('H2:H1000')

    # Save workbook
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


@router.post(
    "/bulk-update",
    responses={
        400: {"description": "File is empty or invalid."},
    },
)
async def update_products_bulk(
    file: Annotated[UploadFile, File(...)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    rows = await _read_file_rows(file)

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or invalid")

    errors: List[str] = []
    updated_count = 0

    for index, row in enumerate(rows, start=2):
        # -------------------------
        # REQUIRED FIELDS
        # -------------------------
        sku = clean_str_with_strip(row.get("product identifier / sku", ""))
        row_type = clean_str_with_strip(row.get("type", ""))

        # -------------------------
        # OPTIONAL ACTION (DEFAULT UPDATE)
        # -------------------------
        raw_action = row.get("action")
        action = (
            clean_str_with_strip(raw_action)
            if raw_action not in (None, "")
            else "update"
        )

        if not sku:
            errors.append(f"Row {index}: SKU missing")
            continue

        if row_type not in ("product", "variant"):
            errors.append(f"Row {index}: Invalid type '{row_type}'")
            continue

        if action not in ("update", "delete"):
            errors.append(f"Row {index}: Invalid action '{action}'")
            continue

        # -------------------------
        # FETCH TARGET
        # -------------------------
        if row_type == "product":
            result = await db.execute(select(Product).where(Product.sku == sku))
            target = result.scalar_one_or_none()
        else:
            result = await db.execute(
                select(ProductVariant).where(ProductVariant.sku == sku)
            )
            target = result.scalar_one_or_none()

        if not target:
            errors.append(f"Row {index}: SKU '{sku}' not found")
            continue

        # -------------------------
        # DELETE
        # -------------------------
        if action == "delete":
            if row_type == "product":
                target.status = "inactive"
            else:
                await db.delete(target)

            updated_count += 1
            continue

        # -------------------------
        # UPDATE (REPLACE)
        # -------------------------
        try:
            if row.get("price") not in (None, ""):
                target.price = float(row["price"])

            if row.get("cost_price") not in (None, ""):
                target.cost_price = float(row["cost_price"])

            if row.get("rrp_price") not in (None, ""):
                target.rrp_price = float(row["rrp_price"])

            if row.get("stock") not in (None, ""):
                target.stock = int(row["stock"])

            updated_count += 1

        except HTTPException as e:
            errors.append(f"Row {index}: {e.detail}")

        except Exception as e:
            errors.append(f"Row {index}: Invalid value ({str(e)})")

    await db.commit()

    return {
        "updated": updated_count,
        "errors": errors,
    }


@router.post(
    "/bulk-update-template",
    responses={
        404: {"description": "No products found."},
    },
)
async def download_bulk_template(
    payload: BulkTemplateRequest,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    product_ids = payload.product_ids
    stmt = (
        select(Product)
        .where(Product.id.in_(product_ids))
        .options(selectinload(Product.variants))
    )

    result = await db.execute(stmt)
    products = result.scalars().all()

    if not products:
        raise HTTPException(status_code=404, detail="No products found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Products Bulk Update"

    # Header
    headers = [
        "product identifier / sku",
        "type",
        "price",
        "cost_price",
        "rrp_price",
        "stock",
        "action",
    ]
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    # Fill rows
    for product in products:
        # Product row - ONLY if no variants
        if not product.variants:
            ws.append(
                [
                    product.sku,
                    "product",
                    float(product.price) if product.price is not None else "",
                    float(product.cost_price) if product.cost_price is not None else "",
                    float(product.rrp_price) if product.rrp_price is not None else "",
                    product.stock,
                    "update/delete",
                ]
            )

        # Variant rows
        for variant in product.variants:
            ws.append(
                [
                    variant.sku,
                    "variant",
                    float(variant.price) if variant.price is not None else "",
                    float(variant.cost_price) if variant.cost_price is not None else "",
                    float(variant.rrp_price) if variant.rrp_price else "",
                    variant.stock,
                    "update/delete",
                ]
            )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=bulk_products_update.xlsx"
        },
    )


IMAGES_HEADERS = [
    "Product Identifier / SKU",
    "Update Type",
    "Product Video Link",
    "Main Image URL",
    "Lifestyle Image URL",
    "Image 1 URL",
    "Image 2 URL",
    "Image 3 URL",
    "Image 4 URL",
    "Image 5 URL",
    "Image 6 URL",
    "Image 7 URL",
    "Image 8 URL",
    "Image 9 URL",
    "Image 10 URL",
]


@router.post(
    "/images-export",
    responses={
        400: {"description": "product_ids required."},
    },
)
async def export_product_media(
    payload: dict,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    product_ids = payload.get("product_ids", [])
    if not product_ids:
        raise HTTPException(status_code=400, detail="product_ids required")

    result = await db.execute(
        select(Product)
        .where(Product.id.in_(product_ids))
        .options(selectinload(Product.variants))
    )
    products = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Product Media"
    # bold headers
    ws.append(IMAGES_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # write rows
    for product in products:
        # --------------------
        # PRODUCT ROW
        # --------------------
        ws.append(
            [
                product.sku,  # Product Identifier / SKU
                "product",  # Update Type
                "",  # Product Video Link
                "",  # Main Image URL
                "",  # Lifestyle Image URL
                "",
                "",
                "",
                "",
                "",  # Image 1–5
                "",
                "",
                "",
                "",
                "",  # Image 6–10
            ]
        )

        # --------------------
        # VARIANT ROWS
        # --------------------
        for variant in product.variants:
            ws.append(
                [
                    variant.sku,  # Product Identifier / SKU
                    "variant",  # Update Type
                    "",  # Product Video Link (not used for variants)
                    "",  # Main Image URL
                    "",  # Lifestyle Image URL
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=product_media_template.xlsx"
        },
    )


@router.post("/images-import")
async def import_product_media(
    file: Annotated[UploadFile, File(...)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    wb = load_workbook(file.file)
    ws = wb.active

    errors = []

    # Skip header row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        sku = row[0]
        update_type = row[1]

        if not sku:
            errors.append({"row": row_idx, "error": "Missing SKU"})
            continue

        if not update_type:
            errors.append({"row": row_idx, "error": "Missing update_type"})
            continue

        (
            video_url,
            main_image,
            lifestyle_image,
            img1,
            img2,
            img3,
            img4,
            img5,
            img6,
            img7,
            img8,
            img9,
            img10,
        ) = row[2:]

        gallery_images = [
            img1,
            img2,
            img3,
            img4,
            img5,
            img6,
            img7,
            img8,
            img9,
            img10,
        ]

        # --------------------------
        # URL VALIDATION
        # --------------------------

        all_urls = [
            ("video_url", video_url),
            ("main_image", main_image),
            ("lifestyle_image", lifestyle_image),
        ]

        for idx, url in enumerate(gallery_images, start=1):
            all_urls.append((f"image_{idx}", url))

        invalid_urls = [name for name, url in all_urls if url and not is_valid_url(url)]

        if invalid_urls:
            errors.append(
                {"row": row_idx, "error": f"Invalid URL(s): {', '.join(invalid_urls)}"}
            )
            continue

        try:
            # --------------------------
            # PRODUCT UPDATE
            # --------------------------
            if update_type == "product":

                result = await db.execute(select(Product).where(Product.sku == sku))
                product = result.scalar_one_or_none()

                if not product:
                    errors.append(
                        {
                            "row": row_idx,
                            "error": f"Product SKU '{sku}' not found",
                        }
                    )
                    continue

                await db.execute(
                    delete(ProductImage).where(ProductImage.product_id == product.id)
                )

                image_order = 1

                if video_url:
                    db.add(
                        ProductImage(
                            product_id=product.id,
                            video_url=video_url.strip(),
                            image_order=image_order,
                        )
                    )
                    image_order += 1

                if main_image:
                    img_url = await download_and_upload_image(main_image.strip(), identifier=product.sku)
                    db.add(
                        ProductImage(
                            product_id=product.id,
                            image_url=img_url,
                            is_main=True,
                            image_order=image_order,
                        )
                    )
                    image_order += 1

                if lifestyle_image:
                    img_url = await download_and_upload_image(lifestyle_image.strip(), identifier=product.sku)
                    db.add(
                        ProductImage(
                            product_id=product.id,
                            image_url=img_url,
                            image_order=image_order,
                        )
                    )
                    image_order += 1

                for url in gallery_images:
                    if url:
                        img_url = await download_and_upload_image(url.strip(), identifier=product.sku)
                        db.add(
                            ProductImage(
                                product_id=product.id,
                                image_url=img_url,
                                image_order=image_order,
                            )
                        )
                        image_order += 1

            # --------------------------
            # VARIANT UPDATE
            # --------------------------
            elif update_type == "variant":

                result = await db.execute(
                    select(ProductVariant).where(ProductVariant.sku == sku)
                )
                variant = result.scalar_one_or_none()

                if not variant:
                    errors.append(
                        {
                            "row": row_idx,
                            "error": f"Variant SKU '{sku}' not found",
                        }
                    )
                    continue

                await db.execute(
                    delete(ProductVariantImage).where(
                        ProductVariantImage.variant_id == variant.id
                    )
                )

                image_order = 1

                if main_image:
                    img_url = await download_and_upload_image(main_image.strip(), identifier=variant.sku)
                    db.add(
                        ProductVariantImage(
                            variant_id=variant.id,
                            image_url=img_url,
                            is_main=True,
                            image_order=image_order,
                        )
                    )
                    image_order += 1

                for url in gallery_images:
                    if url:
                        img_url = await download_and_upload_image(url.strip(), identifier=variant.sku)
                        db.add(
                            ProductVariantImage(
                                variant_id=variant.id,
                                image_url=img_url,
                                image_order=image_order,
                            )
                        )
                        image_order += 1

            else:
                errors.append(
                    {
                        "row": row_idx,
                        "error": f"Invalid update_type '{update_type}'",
                    }
                )
                continue

        except Exception as e:
            errors.append(
                {
                    "row": row_idx,
                    "error": str(e),
                }
            )
            continue

    await db.commit()

    return {
        "message": "Product media import completed",
        "errors": errors,
    }


@router.post("/export-products")
async def export_products(
    payload: ProductExportRequest,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    print("------START",payload.model_dump())
    if payload.background:
        import uuid
        task_id = str(uuid.uuid4())
        
        bg_task = BackgroundTask(
            task_id=task_id,
            task_type="EXPORT_PRODUCTS",
            status="PENDING",
            task_info={"filters": payload.filters} if payload.filters else None,
            added_by="system",  # or current user
        )
        db.add(bg_task)
        await db.commit()
        
        task = export_products_task.apply_async(args=[payload.model_dump()], task_id=task_id)
        print(task.id, "task id******")
        
        return {"task_id": task.id, "status": "PROCESSING"}

    # direct export
    
    headers, rows, timestamp = await build_export_data(
        db,
        payload.product_ids,
        filters=payload.filters,
        sort_key=payload.sort,
        columns=payload.columns,
    )
    stream, filename, media = write_file_for_response(
        headers, rows, timestamp, payload.download_flag
    )

    return StreamingResponse(
        stream,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

